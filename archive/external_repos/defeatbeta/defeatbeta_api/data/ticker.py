import logging
from collections import defaultdict
from decimal import Decimal
from typing import Optional, List, Dict

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.workbook import Workbook

from defeatbeta_api.client.duckdb_client import get_duckdb_client
from defeatbeta_api.client.duckdb_conf import Configuration
from defeatbeta_api.client.hugging_face_client import HuggingFaceClient
from defeatbeta_api.data.balance_sheet import BalanceSheet
from defeatbeta_api.data.finance_item import FinanceItem
from defeatbeta_api.data.finance_value import FinanceValue
from defeatbeta_api.data.income_statement import IncomeStatement
from defeatbeta_api.data.news import News
from defeatbeta_api.data.print_visitor import PrintVisitor
from defeatbeta_api.data.sql.sql_loader import load_sql
from defeatbeta_api.data.statement import Statement
from defeatbeta_api.data.stock_statement import StockStatement
from defeatbeta_api.data.transcripts import Transcripts
from defeatbeta_api.data.treasure import Treasure
from defeatbeta_api.data.company_meta import CompanyMeta
from defeatbeta_api.utils.case_insensitive_dict import CaseInsensitiveDict
from defeatbeta_api.utils.const import stock_profile, stock_earning_calendar, stock_officers, \
    stock_split_events, \
    stock_dividend_events, stock_tailing_eps, \
    stock_prices, stock_statement, income_statement, balance_sheet, cash_flow, quarterly, annual, \
    stock_earning_call_transcripts, stock_news, stock_revenue_breakdown, stock_shares_outstanding, exchange_rate, \
    stock_sec_filing
from defeatbeta_api.utils.util import load_finance_template, parse_all_title_keys, income_statement_template_type, \
    balance_sheet_template_type, cash_flow_template_type, sp500_cagr_returns_rolling, validate_dcf_directory, \
    in_notebook


class Ticker:
    def __init__(self, ticker, http_proxy: Optional[str] = None, log_level: Optional[str] = logging.INFO, config: Optional[Configuration] = None):
        self.ticker = ticker.upper()
        self.http_proxy = http_proxy
        self.config = config
        self.duckdb_client = get_duckdb_client(http_proxy=self.http_proxy, log_level=log_level, config=config)
        self.huggingface_client = HuggingFaceClient()
        self.log_level = log_level
        self.treasure = Treasure(
            http_proxy=self.http_proxy,
            log_level=self.log_level,
            config=config
        )
        self.company_meta = CompanyMeta(
            http_proxy=self.http_proxy,
            log_level=self.log_level,
            config=config
        )

    def info(self) -> pd.DataFrame:
        return self._query_data(stock_profile)

    def sec_filing(self) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_sec_filing)
        sql = load_sql("select_sec_filing_by_symbol", ticker=self.ticker, url=url)
        return self.duckdb_client.query(sql)

    def officers(self) -> pd.DataFrame:
        return self._query_data(stock_officers)

    def calendar(self) -> pd.DataFrame:
        return self._query_data(stock_earning_calendar)

    def splits(self) -> pd.DataFrame:
        return self._query_data(stock_split_events)

    def dividends(self) -> pd.DataFrame:
        return self._query_data(stock_dividend_events)

    def ttm_eps(self) -> pd.DataFrame:
        return self._query_data(stock_tailing_eps)

    def price(self) -> pd.DataFrame:
        return self._query_data(stock_prices)

    def beta(self, period: str = "5y", benchmark: str = "SPY") -> pd.DataFrame:
        """
        Calculate beta for the stock relative to a benchmark index.
        Uses monthly returns for periods >= 1 year (standard industry practice).

        Args:
            period: Time period in format like "30d", "3m", "1y", "5y"
                   d=days, m=months, y=years
            benchmark: Benchmark symbol (default: SPY for S&P 500)

        Returns:
            DataFrame with columns: symbol, report_date, beta, period, benchmark

        Example:
            ticker = Ticker("AAPL")
            df = ticker.beta("1y")  # 1-year beta (12 monthly returns)
            df = ticker.beta("5y")  # 5-year beta (60 monthly returns)
        """
        from datetime import datetime, timedelta
        import re

        # Parse period string
        match = re.match(r'^(\d+)([dmy])$', period.lower())
        if not match:
            raise ValueError(f"Invalid period format: {period}. Use format like '30d', '3m', '1y'")

        value, unit = int(match.group(1)), match.group(2)

        # Use data update time as end date (data may not be current to today)
        end_date = datetime.strptime(self.huggingface_client.get_data_update_time(), '%Y-%m-%d')
        if unit == 'd':
            start_date = end_date - timedelta(days=value)
        elif unit == 'm':
            start_date = end_date - timedelta(days=value * 30)
        elif unit == 'y':
            start_date = end_date - timedelta(days=value * 365)

        # Get price data for stock and benchmark using SQL file
        url = self.huggingface_client.get_url_path(stock_prices)
        sql = load_sql(
            "select_beta_prices_by_symbol",
            ticker=self.ticker,
            benchmark=benchmark,
            url=url,
            start_date=start_date.strftime('%Y-%m-%d'),
            end_date=end_date.strftime('%Y-%m-%d')
        )
        merged_df = self.duckdb_client.query(sql)

        if len(merged_df) < 2:
            raise ValueError(f"Insufficient data for period {period}")

        # Convert report_date to datetime
        merged_df['report_date'] = pd.to_datetime(merged_df['report_date']).astype('datetime64[us]')

        # For periods >= 1 year, use monthly returns (industry standard)
        # For shorter periods, use daily returns
        if (unit == 'y') or (unit == 'm' and value >= 12):
            # Resample to month-end and take last closing price of each month
            merged_df = merged_df.set_index('report_date')
            monthly_df = merged_df.resample('ME').last().dropna()

            # Calculate monthly returns
            monthly_df['stock_return'] = monthly_df['stock_close'].pct_change()
            monthly_df['benchmark_return'] = monthly_df['benchmark_close'].pct_change()

            # Drop first row (NaN returns)
            monthly_df = monthly_df.dropna()

            if len(monthly_df) < 2:
                raise ValueError(f"Insufficient monthly data for period {period}")

            # Calculate beta using covariance and variance
            covariance = np.cov(monthly_df['stock_return'], monthly_df['benchmark_return'])[0, 1]
            benchmark_variance = np.var(monthly_df['benchmark_return'], ddof=1)
        else:
            # Use daily returns for short periods
            merged_df['stock_return'] = merged_df['stock_close'].pct_change()
            merged_df['benchmark_return'] = merged_df['benchmark_close'].pct_change()

            # Drop first row (NaN returns)
            merged_df = merged_df.dropna()

            # Calculate beta using covariance and variance
            covariance = np.cov(merged_df['stock_return'], merged_df['benchmark_return'])[0, 1]
            benchmark_variance = np.var(merged_df['benchmark_return'], ddof=1)

        beta = covariance / benchmark_variance

        return pd.DataFrame([{
            'symbol': self.ticker,
            'report_date': end_date.strftime('%Y-%m-%d'),
            'beta': round(beta, 4),
            'period': period,
            'benchmark': benchmark,
        }])

    def currency(self, symbol: str) -> pd.DataFrame:
        df = self._query_data2(exchange_rate, symbol)
        df['report_date'] = pd.to_datetime(df['report_date']).astype('datetime64[us]')
        return df

    def shares(self) -> pd.DataFrame:
        return self._query_data(stock_shares_outstanding)

    def quarterly_income_statement(self) -> Statement:
        return self._statement(income_statement, quarterly)

    def annual_income_statement(self) -> Statement:
        return self._statement(income_statement, annual)

    def quarterly_balance_sheet(self) -> Statement:
        return self._statement(balance_sheet, quarterly)

    def annual_balance_sheet(self) -> Statement:
        return self._statement(balance_sheet, annual)

    def quarterly_cash_flow(self) -> Statement:
        return self._statement(cash_flow, quarterly)

    def annual_cash_flow(self) -> Statement:
        return self._statement(cash_flow, annual)

    def ttm_pe(self) -> pd.DataFrame:
        price_df = self.price()

        eps_df = self.ttm_eps()

        price_df['report_date'] = pd.to_datetime(price_df['report_date']).astype('datetime64[us]')
        eps_df['report_date'] = pd.to_datetime(eps_df['report_date']).astype('datetime64[us]')

        result_df = price_df.copy()
        result_df = result_df.rename(columns={'report_date': 'price_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('price_report_date'),
            eps_df.sort_values('report_date'),
            left_on='price_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df['ttm_pe'] = round(result_df['close'] / result_df['tailing_eps'], 2)
        # Negative EPS yields a meaningless negative P/E; mask per Bloomberg/FactSet convention
        result_df.loc[result_df['tailing_eps'] < 0, 'ttm_pe'] = float('nan')

        result_df = result_df[[
            'price_report_date',
            'report_date',
            'close',
            'tailing_eps',
            'ttm_pe'
        ]]

        result_df = result_df.rename(columns={
            'price_report_date': 'report_date',
            'close': 'close_price',
            'tailing_eps': 'ttm_eps',
            'report_date': 'eps_report_date'
        })

        result_df.insert(0, 'symbol', self.ticker)
        result_df = result_df.dropna(subset=['ttm_eps']).reset_index(drop=True)
        return result_df

    def quarterly_gross_margin(self) -> pd.DataFrame:
        return self._generate_margin('gross', 'quarterly', 'gross_profit', 'gross_margin')

    def annual_gross_margin(self) -> pd.DataFrame:
        return self._generate_margin('gross', 'annual', 'gross_profit', 'gross_margin')

    def quarterly_operating_margin(self) -> pd.DataFrame:
        return self._generate_margin('operating', 'quarterly', 'operating_income', 'operating_margin')

    def annual_operating_margin(self) -> pd.DataFrame:
        return self._generate_margin('operating', 'annual', 'operating_income', 'operating_margin')

    def quarterly_net_margin(self) -> pd.DataFrame:
        return self._generate_margin('net', 'quarterly', 'net_income_common_stockholders', 'net_margin')

    def annual_net_margin(self) -> pd.DataFrame:
        return self._generate_margin('net', 'annual', 'net_income_common_stockholders', 'net_margin')

    def quarterly_ebitda_margin(self) -> pd.DataFrame:
        return self._generate_margin('ebitda', 'quarterly', 'ebitda', 'ebitda_margin')

    def annual_ebitda_margin(self) -> pd.DataFrame:
        return self._generate_margin('ebitda', 'annual', 'ebitda', 'ebitda_margin')

    def quarterly_fcf_margin(self) -> pd.DataFrame:
        return self._generate_margin('fcf', 'quarterly', 'free_cash_flow', 'fcf_margin')

    def annual_fcf_margin(self) -> pd.DataFrame:
        return self._generate_margin('fcf', 'annual', 'free_cash_flow', 'fcf_margin')

    def earning_call_transcripts(self) -> Transcripts:
        return Transcripts(self.ticker, self._query_data(stock_earning_call_transcripts), self.log_level)

    def news(self) -> News:
        url = self.huggingface_client.get_url_path(stock_news)
        sql = load_sql("select_news_by_symbol", ticker = self.ticker, url = url)
        return News(self.duckdb_client.query(sql))

    def revenue_by_segment(self) -> pd.DataFrame:
        return self._revenue_by_breakdown('segment')

    def revenue_by_geography(self) -> pd.DataFrame:
        return self._revenue_by_breakdown('geography')

    def revenue_by_product(self) -> pd.DataFrame:
        return self._revenue_by_breakdown('product')

    def quarterly_revenue_yoy_growth(self) -> pd.DataFrame:
        return self._calculate_yoy_growth(item_name='total_revenue', period_type='quarterly', finance_type='income_statement')

    def annual_revenue_yoy_growth(self) -> pd.DataFrame:
        return self._calculate_yoy_growth(item_name='total_revenue', period_type='annual', finance_type='income_statement')

    def quarterly_operating_income_yoy_growth(self) -> pd.DataFrame:
        return self._calculate_yoy_growth(item_name='operating_income', period_type='quarterly', finance_type='income_statement')

    def annual_operating_income_yoy_growth(self) -> pd.DataFrame:
        return self._calculate_yoy_growth(item_name='operating_income', period_type='annual', finance_type='income_statement')

    def quarterly_ebitda_yoy_growth(self) -> pd.DataFrame:
        return self._calculate_yoy_growth(item_name='ebitda', period_type='quarterly', finance_type='income_statement')

    def annual_ebitda_yoy_growth(self) -> pd.DataFrame:
        return self._calculate_yoy_growth(item_name='ebitda', period_type='annual', finance_type='income_statement')

    def quarterly_net_income_yoy_growth(self) -> pd.DataFrame:
        return self._calculate_yoy_growth(item_name='net_income_common_stockholders', period_type='quarterly', finance_type='income_statement')

    def annual_net_income_yoy_growth(self) -> pd.DataFrame:
        return self._calculate_yoy_growth(item_name='net_income_common_stockholders', period_type='annual', finance_type='income_statement')

    def quarterly_fcf_yoy_growth(self) -> pd.DataFrame:
        return self._calculate_yoy_growth(item_name='free_cash_flow', period_type='quarterly', finance_type='cash_flow')

    def annual_fcf_yoy_growth(self) -> pd.DataFrame:
        return self._calculate_yoy_growth(item_name='free_cash_flow', period_type='annual', finance_type='cash_flow')

    def quarterly_eps_yoy_growth(self) -> pd.DataFrame:
        return self._quarterly_eps_yoy_growth('eps', 'eps', 'prev_year_eps')

    def quarterly_ttm_eps_yoy_growth(self) -> pd.DataFrame:
        return self._quarterly_eps_yoy_growth('tailing_eps', 'ttm_eps', 'prev_year_ttm_eps')

    def market_capitalization(self) -> pd.DataFrame:
        price_df = self.price()

        shares_df = self.shares()

        price_df['report_date'] = pd.to_datetime(price_df['report_date']).astype('datetime64[us]')
        shares_df['report_date'] = pd.to_datetime(shares_df['report_date']).astype('datetime64[us]')

        result_df = price_df.copy()
        result_df = result_df.rename(columns={'report_date': 'price_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('price_report_date'),
            shares_df.sort_values('report_date'),
            left_on='price_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df['market_cap'] = round(result_df['close'] * result_df['shares_outstanding'], 2)

        result_df = result_df[[
            'price_report_date',
            'report_date',
            'close',
            'shares_outstanding',
            'market_cap'
        ]]

        result_df = result_df.rename(columns={
            'price_report_date': 'report_date',
            'close': 'close_price',
            'report_date': 'shares_report_date',
            'market_cap': 'market_capitalization'
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def ps_ratio(self) -> pd.DataFrame:
        market_cap_df = self.market_capitalization()
        ttm_revenue_df = self.ttm_revenue()

        market_cap_df['report_date'] = pd.to_datetime(market_cap_df['report_date']).astype('datetime64[us]')
        ttm_revenue_df['report_date'] = pd.to_datetime(ttm_revenue_df['report_date']).astype('datetime64[us]')

        result_df = market_cap_df.copy()
        result_df = result_df.rename(columns={'report_date': 'market_cap_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('market_cap_report_date'),
            ttm_revenue_df.sort_values('report_date'),
            left_on='market_cap_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df = result_df[result_df['report_date'].notna()]

        result_df['ps_ratio'] = round(result_df['market_capitalization'] / result_df['ttm_total_revenue_usd'], 2)

        result_df = result_df[[
            'market_cap_report_date',
            'market_capitalization',
            'report_date',
            'ttm_total_revenue',
            'exchange_to_usd_rate',
            'ttm_total_revenue_usd',
            'ps_ratio'
        ]]

        result_df = result_df.rename(columns={
            'market_cap_report_date': 'report_date',
            'report_date': 'fiscal_quarter',
            'ttm_total_revenue': 'ttm_revenue',
            'exchange_to_usd_rate': 'exchange_rate',
            'ttm_total_revenue_usd': 'ttm_revenue_usd'
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def pb_ratio(self) -> pd.DataFrame:
        market_cap_df = self.market_capitalization()
        bve_df = self._quarterly_book_value_of_equity()

        market_cap_df['report_date'] = pd.to_datetime(market_cap_df['report_date']).astype('datetime64[us]')
        bve_df['report_date'] = pd.to_datetime(bve_df['report_date']).astype('datetime64[us]')

        result_df = market_cap_df.copy()
        result_df = result_df.rename(columns={'report_date': 'market_cap_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('market_cap_report_date'),
            bve_df.sort_values('report_date'),
            left_on='market_cap_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df = result_df[result_df['report_date'].notna()]

        result_df['pb_ratio'] = round(result_df['market_capitalization'] / result_df['book_value_of_equity_usd'], 2)

        result_df = result_df[[
            'market_cap_report_date',
            'market_capitalization',
            'report_date',
            'book_value_of_equity',
            'exchange_to_usd_rate',
            'book_value_of_equity_usd',
            'pb_ratio'
        ]]

        result_df = result_df.rename(columns={
            'market_cap_report_date': 'report_date',
            'report_date': 'fiscal_quarter',
            'ttm_total_revenue': 'book_value_of_equity',
            'exchange_to_usd_rate': 'exchange_rate',
            'ttm_total_revenue_usd': 'book_value_of_equity_usd'
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def debt_to_equity(self) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_statement)
        sql = load_sql("select_debt_to_equity_by_symbol", ticker=self.ticker, url=url)
        result_df = self.duckdb_client.query(sql)

        result_df['stockholders_equity'] = result_df['stockholders_equity'].fillna(0)

        result_df['debt_to_equity'] = (result_df['total_debt'] / result_df['stockholders_equity']).replace([np.inf, -np.inf], np.nan).round(2)

        return result_df

    def net_debt_ttm(self) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_statement)
        sql = load_sql("select_net_debt_ttm_by_symbol", ticker=self.ticker, url=url)
        result_df = self.duckdb_client.query(sql)
        result_df = result_df[[
            'symbol',
            'report_date',
            'long_term_debt',
            'cash_and_short_term_investments',
            'net_debt',
            'avg_net_debt_ttm'
        ]]
        return result_df

    def enterprise_value(self) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_statement)
        sql = load_sql("select_enterprise_value_components_by_symbol", ticker=self.ticker, url=url)
        ev_components_df = self.duckdb_client.query(sql)

        company_info = self.company_meta.get_company_info(self.ticker)
        currency = company_info["financial_currency"] if company_info and company_info.get("financial_currency") else 'USD'

        if currency == 'USD':
            currency_df = pd.DataFrame()
            currency_df['report_date'] = pd.to_datetime(ev_components_df['report_date']).astype('datetime64[us]')
            currency_df['close'] = 1.0
        else:
            currency_df = self.currency(currency + '=X')

        ev_components_df['report_date'] = pd.to_datetime(ev_components_df['report_date']).astype('datetime64[us]')
        currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

        ev_components_df = pd.merge_asof(
            ev_components_df.sort_values('report_date'),
            currency_df[['report_date', 'close']].sort_values('report_date'),
            on='report_date',
            direction='backward'
        )
        ev_components_df = ev_components_df.rename(columns={'close': 'exchange_to_usd_rate'})

        for col in ['total_debt', 'minority_interest', 'preferred_stock_equity', 'cash_and_cash_equivalents']:
            ev_components_df[col] = ev_components_df[col].fillna(0)
            ev_components_df[f'{col}_usd'] = round(ev_components_df[col] / ev_components_df['exchange_to_usd_rate'], 2)

        market_cap_df = self.market_capitalization().drop(columns=['symbol'])
        market_cap_df['report_date'] = pd.to_datetime(market_cap_df['report_date']).astype('datetime64[us]')

        result_df = market_cap_df.copy()
        result_df = result_df.rename(columns={'report_date': 'market_cap_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('market_cap_report_date'),
            ev_components_df.sort_values('report_date'),
            left_on='market_cap_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df = result_df[result_df['report_date'].notna()]

        result_df['enterprise_value'] = round(
            result_df['market_capitalization']
            + result_df['total_debt_usd']
            + result_df['minority_interest_usd']
            + result_df['preferred_stock_equity_usd']
            - result_df['cash_and_cash_equivalents_usd'],
            2
        )

        result_df = result_df[[
            'market_cap_report_date',
            'market_capitalization',
            'report_date',
            'exchange_to_usd_rate',
            'total_debt',
            'total_debt_usd',
            'minority_interest',
            'minority_interest_usd',
            'preferred_stock_equity',
            'preferred_stock_equity_usd',
            'cash_and_cash_equivalents',
            'cash_and_cash_equivalents_usd',
            'enterprise_value'
        ]]

        result_df = result_df.rename(columns={
            'market_cap_report_date': 'report_date',
            'report_date': 'fiscal_quarter',
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def enterprise_to_revenue(self) -> pd.DataFrame:
        ev_df = self.enterprise_value().drop(columns=['symbol'])
        ttm_revenue_df = self.ttm_revenue().drop(columns=['symbol'])

        ev_df['report_date'] = pd.to_datetime(ev_df['report_date']).astype('datetime64[us]')
        ttm_revenue_df['report_date'] = pd.to_datetime(ttm_revenue_df['report_date']).astype('datetime64[us]')

        result_df = ev_df.copy()
        result_df = result_df.rename(columns={'report_date': 'ev_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('ev_report_date'),
            ttm_revenue_df[['report_date', 'ttm_total_revenue', 'ttm_total_revenue_usd']].sort_values('report_date'),
            left_on='ev_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df = result_df[result_df['report_date'].notna()]

        result_df['ev_to_revenue'] = (result_df['enterprise_value'] / result_df['ttm_total_revenue_usd']).replace([np.inf, -np.inf], np.nan).round(2)

        result_df = result_df[[
            'ev_report_date',
            'enterprise_value',
            'report_date',
            'ttm_total_revenue',
            'ttm_total_revenue_usd',
            'ev_to_revenue'
        ]]

        result_df = result_df.rename(columns={
            'ev_report_date': 'report_date',
            'report_date': 'fiscal_quarter',
            'ttm_total_revenue': 'ttm_revenue',
            'ttm_total_revenue_usd': 'ttm_revenue_usd'
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def peg_ratio(self) -> pd.DataFrame:
        ttm_pe_df = self.ttm_pe()
        eps_yoy_df = self.quarterly_eps_yoy_growth()

        ttm_pe_df['report_date'] = pd.to_datetime(ttm_pe_df['report_date']).astype('datetime64[us]')
        ttm_pe_df['eps_report_date'] = pd.to_datetime(ttm_pe_df['eps_report_date']).astype('datetime64[us]')
        eps_yoy_df['report_date'] = pd.to_datetime(eps_yoy_df['report_date']).astype('datetime64[us]')

        result_df = ttm_pe_df.copy()
        result_df = result_df.rename(columns={'report_date': 'ttm_pe_report_date'})
        result_df = result_df[result_df['eps_report_date'].notna()]

        result_df = pd.merge_asof(
            result_df,
            eps_yoy_df,
            left_on='eps_report_date',
            right_on='report_date',
            direction='backward'
        )

        valid_peg = (result_df['ttm_eps'] > 0) & (result_df['yoy_growth'] > 0)
        result_df['peg_ratio'] = np.where(
            valid_peg,
            (result_df['ttm_pe'] / (result_df['yoy_growth'] * 100)).round(2),
            np.nan
        )

        result_df = result_df[[
            'ttm_pe_report_date',
            'close_price',
            'report_date',
            'ttm_eps',
            'ttm_pe',
            'yoy_growth',
            'peg_ratio'
        ]]

        result_df = result_df.rename(columns={
            'ttm_pe_report_date': 'report_date',
            'report_date': 'fiscal_quarter',
            'yoy_growth': 'eps_yoy_growth'
        })

        result_df = result_df[result_df['ttm_pe'].notna()]
        result_df = result_df[result_df['eps_yoy_growth'].notna()]
        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def _quarterly_book_value_of_equity(self) -> pd.DataFrame:
        stockholders_equity_url = self.huggingface_client.get_url_path(stock_statement)
        stockholders_equity_sql = load_sql("select_quarterly_book_value_of_equity_by_symbol",
                                           ticker = self.ticker,
                                           stockholders_equity_url = stockholders_equity_url)
        stockholders_equity_df = self.duckdb_client.query(stockholders_equity_sql)

        company_info = self.company_meta.get_company_info(self.ticker)
        currency = company_info["financial_currency"] if company_info and company_info.get("financial_currency") else 'USD'

        if currency == 'USD':
            currency_df = pd.DataFrame()
            currency_df['report_date'] = pd.to_datetime(
                stockholders_equity_df['report_date']).astype('datetime64[us]')
            currency_df['symbol'] = currency + '=X'
            currency_df['open'] = 1.0
            currency_df['close'] = 1.0
            currency_df['high'] = 1.0
            currency_df['low'] = 1.0
        else:
            currency_df = self.currency(currency + '=X')

        stockholders_equity_df['report_date'] = pd.to_datetime(stockholders_equity_df['report_date']).astype('datetime64[us]')
        currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

        result_df = stockholders_equity_df.copy()
        result_df = result_df.rename(columns={'report_date': 'book_value_of_equity_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('book_value_of_equity_report_date'),
            currency_df.sort_values('report_date'),
            left_on='book_value_of_equity_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df['book_value_of_equity_usd'] = round(result_df['book_value_of_equity'] / result_df['close'], 2)

        result_df = result_df[[
            'book_value_of_equity_report_date',
            'book_value_of_equity',
            'report_date',
            'close',
            'book_value_of_equity_usd'
        ]]

        result_df = result_df.rename(columns={
            'book_value_of_equity_report_date': 'report_date',
            'report_date': 'exchange_report_date',
            'close': 'exchange_to_usd_rate'
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def ttm_revenue(self) -> pd.DataFrame:
        ttm_revenue_url = self.huggingface_client.get_url_path(stock_statement)
        ttm_revenue_sql = load_sql("select_ttm_revenue_by_symbol",
                                   ticker = self.ticker,
                                   ttm_revenue_url = ttm_revenue_url)
        ttm_revenue_df = self.duckdb_client.query(ttm_revenue_sql)

        company_info = self.company_meta.get_company_info(self.ticker)
        currency = company_info["financial_currency"] if company_info and company_info.get("financial_currency") else 'USD'
        if currency == 'USD':
            currency_df = pd.DataFrame()
            currency_df['report_date'] = pd.to_datetime(
                ttm_revenue_df['report_date']).astype('datetime64[us]')
            currency_df['symbol'] = currency + '=X'
            currency_df['open'] = 1.0
            currency_df['close'] = 1.0
            currency_df['high'] = 1.0
            currency_df['low'] = 1.0
        else:
            currency_df = self.currency(symbol = currency + '=X')

        ttm_revenue_df['report_date'] = pd.to_datetime(ttm_revenue_df['report_date']).astype('datetime64[us]')
        currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

        result_df = ttm_revenue_df.copy()
        result_df = result_df.rename(columns={'report_date': 'ttm_revenue_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('ttm_revenue_report_date'),
            currency_df.sort_values('report_date'),
            left_on='ttm_revenue_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df['ttm_total_revenue_usd'] = round(result_df['ttm_total_revenue'] / result_df['close'], 2)

        result_df = result_df[[
            'ttm_revenue_report_date',
            'ttm_total_revenue',
            'report_date_2_revenue',
            'report_date',
            'close',
            'ttm_total_revenue_usd'
        ]]

        result_df = result_df.rename(columns={
            'ttm_revenue_report_date': 'report_date',
            'report_date': 'exchange_report_date',
            'close': 'exchange_to_usd_rate'
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def ttm_fcf(self) -> pd.DataFrame:
        ttm_fcf_url = self.huggingface_client.get_url_path(stock_statement)
        ttm_fcf_sql = load_sql("select_ttm_fcf_by_symbol",
                              ticker=self.ticker,
                              ttm_fcf_url=ttm_fcf_url)
        ttm_fcf_df = self.duckdb_client.query(ttm_fcf_sql)

        company_info = self.company_meta.get_company_info(self.ticker)
        currency = company_info["financial_currency"] if company_info and company_info.get("financial_currency") else 'USD'
        if currency == 'USD':
            currency_df = pd.DataFrame()
            currency_df['report_date'] = pd.to_datetime(
                ttm_fcf_df['report_date']).astype('datetime64[us]')
            currency_df['symbol'] = currency + '=X'
            currency_df['open'] = 1.0
            currency_df['close'] = 1.0
            currency_df['high'] = 1.0
            currency_df['low'] = 1.0
        else:
            currency_df = self.currency(symbol=currency + '=X')

        ttm_fcf_df['report_date'] = pd.to_datetime(ttm_fcf_df['report_date']).astype('datetime64[us]')
        currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

        result_df = ttm_fcf_df.copy()
        result_df = result_df.rename(columns={'report_date': 'ttm_fcf_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('ttm_fcf_report_date'),
            currency_df.sort_values('report_date'),
            left_on='ttm_fcf_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df['ttm_free_cash_flow_usd'] = round(result_df['ttm_free_cash_flow'] / result_df['close'], 2)

        result_df = result_df[[
            'ttm_fcf_report_date',
            'ttm_free_cash_flow',
            'report_date_2_fcf',
            'report_date',
            'close',
            'ttm_free_cash_flow_usd'
        ]]

        result_df = result_df.rename(columns={
            'ttm_fcf_report_date': 'report_date',
            'report_date': 'exchange_report_date',
            'close': 'exchange_to_usd_rate'
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def ttm_ebitda(self) -> pd.DataFrame:
        ttm_ebitda_url = self.huggingface_client.get_url_path(stock_statement)
        ttm_ebitda_sql = load_sql("select_ttm_ebitda_by_symbol",
                                  ticker=self.ticker,
                                  ttm_ebitda_url=ttm_ebitda_url)
        ttm_ebitda_df = self.duckdb_client.query(ttm_ebitda_sql)

        company_info = self.company_meta.get_company_info(self.ticker)
        currency = company_info["financial_currency"] if company_info and company_info.get("financial_currency") else 'USD'
        if currency == 'USD':
            currency_df = pd.DataFrame()
            currency_df['report_date'] = pd.to_datetime(ttm_ebitda_df['report_date']).astype('datetime64[us]')
            currency_df['symbol'] = currency + '=X'
            currency_df['open'] = 1.0
            currency_df['close'] = 1.0
            currency_df['high'] = 1.0
            currency_df['low'] = 1.0
        else:
            currency_df = self.currency(symbol=currency + '=X')

        ttm_ebitda_df['report_date'] = pd.to_datetime(ttm_ebitda_df['report_date']).astype('datetime64[us]')
        currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

        result_df = ttm_ebitda_df.copy()
        result_df = result_df.rename(columns={'report_date': 'ttm_ebitda_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('ttm_ebitda_report_date'),
            currency_df.sort_values('report_date'),
            left_on='ttm_ebitda_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df['ttm_ebitda_usd'] = round(result_df['ttm_ebitda'] / result_df['close'], 2)

        result_df = result_df[[
            'ttm_ebitda_report_date',
            'ttm_ebitda',
            'report_date_2_ebitda',
            'report_date',
            'close',
            'ttm_ebitda_usd'
        ]]

        result_df = result_df.rename(columns={
            'ttm_ebitda_report_date': 'report_date',
            'report_date': 'exchange_report_date',
            'close': 'exchange_to_usd_rate'
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def enterprise_to_ebitda(self) -> pd.DataFrame:
        ev_df = self.enterprise_value().drop(columns=['symbol'])
        ttm_ebitda_df = self.ttm_ebitda().drop(columns=['symbol'])

        ev_df['report_date'] = pd.to_datetime(ev_df['report_date']).astype('datetime64[us]')
        ttm_ebitda_df['report_date'] = pd.to_datetime(ttm_ebitda_df['report_date']).astype('datetime64[us]')

        result_df = ev_df.copy()
        result_df = result_df.rename(columns={'report_date': 'ev_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('ev_report_date'),
            ttm_ebitda_df[['report_date', 'ttm_ebitda', 'ttm_ebitda_usd']].sort_values('report_date'),
            left_on='ev_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df = result_df[result_df['report_date'].notna()]

        result_df['ev_to_ebitda'] = (result_df['enterprise_value'] / result_df['ttm_ebitda_usd']).replace([np.inf, -np.inf], np.nan).round(2)

        result_df = result_df[[
            'ev_report_date',
            'enterprise_value',
            'report_date',
            'ttm_ebitda',
            'ttm_ebitda_usd',
            'ev_to_ebitda'
        ]]

        result_df = result_df.rename(columns={
            'ev_report_date': 'report_date',
            'report_date': 'fiscal_quarter',
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def ttm_net_income_common_stockholders(self) -> pd.DataFrame:
        ttm_net_income_url = self.huggingface_client.get_url_path(stock_statement)
        ttm_net_income_sql = load_sql("select_ttm_net_income_common_stockholders_by_symbol",
                                      ticker=self.ticker,
                                      ttm_net_income_url=ttm_net_income_url)
        ttm_net_income_df = self.duckdb_client.query(ttm_net_income_sql)

        company_info = self.company_meta.get_company_info(self.ticker)
        currency = company_info["financial_currency"] if company_info and company_info.get("financial_currency") else 'USD'

        if currency == 'USD':
            currency_df = pd.DataFrame()
            currency_df['report_date'] = pd.to_datetime(
                ttm_net_income_df['report_date']).astype('datetime64[us]')
            currency_df['symbol'] = currency + '=X'
            currency_df['open'] = 1.0
            currency_df['close'] = 1.0
            currency_df['high'] = 1.0
            currency_df['low'] = 1.0
        else:
            currency_df = self.currency(symbol = currency + '=X')

        ttm_net_income_df['report_date'] = pd.to_datetime(ttm_net_income_df['report_date']).astype('datetime64[us]')
        currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

        result_df = ttm_net_income_df.copy()
        result_df = result_df.rename(columns={'report_date': 'ttm_net_income_report_date'})

        result_df = pd.merge_asof(
            result_df.sort_values('ttm_net_income_report_date'),
            currency_df.sort_values('report_date'),
            left_on='ttm_net_income_report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df['ttm_net_income_usd'] = round(result_df['ttm_net_income'] / result_df['close'], 2)

        result_df = result_df[[
            'ttm_net_income_report_date',
            'ttm_net_income',
            'report_date_2_net_income',
            'report_date',
            'close',
            'ttm_net_income_usd'
        ]]

        result_df = result_df.rename(columns={
            'ttm_net_income_report_date': 'report_date',
            'report_date': 'exchange_report_date',
            'close': 'exchange_to_usd_rate'
        })

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def roe(self) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_statement)
        sql = load_sql("select_roe_by_symbol", ticker = self.ticker, url = url)
        result_df = self.duckdb_client.query(sql)
        result_df = result_df[[
            'symbol',
            'report_date',
            'net_income_common_stockholders',
            'beginning_stockholders_equity',
            'ending_stockholders_equity',
            'avg_equity',
            'roe'
        ]]
        return result_df

    def roa(self) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_statement)
        sql = load_sql("select_roa_by_symbol", ticker = self.ticker, url = url)
        result_df = self.duckdb_client.query(sql)
        result_df = result_df[[
            'symbol',
            'report_date',
            'net_income_common_stockholders',
            'beginning_total_assets',
            'ending_total_assets',
            'avg_assets',
            'roa'
        ]]
        return result_df

    def roic(self) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_statement)
        sql = load_sql("select_roic_by_symbol", ticker = self.ticker, url = url)
        result_df = self.duckdb_client.query(sql)
        result_df = result_df[[
            'symbol',
            'report_date',
            'ebit',
            'tax_rate_for_calcs',
            'nopat',
            'beginning_invested_capital',
            'ending_invested_capital',
            'avg_invested_capital',
            'roic'
        ]]
        return result_df

    def roce(self) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_statement)
        sql = load_sql("select_roce_by_symbol", ticker=self.ticker, url=url)
        result_df = self.duckdb_client.query(sql)
        result_df = result_df[[
            'symbol',
            'report_date',
            'ebit',
            'beginning_capital_employed',
            'ending_capital_employed',
            'avg_capital_employed',
            'roce'
        ]]
        return result_df

    def equity_multiplier(self) -> pd.DataFrame:
        roe = self.roe()
        roa = self.roa()

        roe['report_date'] = pd.to_datetime(roe['report_date']).astype('datetime64[us]')
        roa['report_date'] = pd.to_datetime(roa['report_date']).astype('datetime64[us]')

        result_df = pd.merge_asof(
            roe,
            roa,
            left_on='report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df['equity_multiplier'] = round(result_df['roe'] / result_df['roa'], 2)

        result_df = result_df[[
            'report_date',
            'roe',
            'roa',
            'equity_multiplier'
        ]]

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def asset_turnover(self) -> pd.DataFrame:
        roa = self.roa()
        quarterly_net_margin = self.quarterly_net_margin()

        roa['report_date'] = pd.to_datetime(roa['report_date']).astype('datetime64[us]')
        quarterly_net_margin['report_date'] = pd.to_datetime(quarterly_net_margin['report_date']).astype('datetime64[us]')

        result_df = pd.merge_asof(
            roa,
            quarterly_net_margin,
            left_on='report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df['asset_turnover'] = round(result_df['roa'] / result_df['net_margin'], 2)

        result_df = result_df[[
            'report_date',
            'roa',
            'net_margin',
            'asset_turnover'
        ]]

        result_df.insert(0, 'symbol', self.ticker)
        return result_df

    def wacc(self) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_statement)
        sql = load_sql("select_wacc_by_symbol", ticker = self.ticker, url = url)
        wacc_df = self.duckdb_client.query(sql)
        company_info = self.company_meta.get_company_info(self.ticker)
        currency = company_info["financial_currency"] if company_info and company_info.get("financial_currency") else 'USD'

        if currency == 'USD':
            currency_df = pd.DataFrame()
            currency_df['report_date'] = pd.to_datetime(
                wacc_df['report_date']).astype('datetime64[us]')
            currency_df['symbol'] = currency + '=X'
            currency_df['open'] = 1.0
            currency_df['close'] = 1.0
            currency_df['high'] = 1.0
            currency_df['low'] = 1.0
        else:
            currency_df = self.currency(symbol = currency + '=X')

        currency_df = currency_df[[
            'report_date',
            'close'
        ]]
        currency_df = currency_df.rename(columns={
            'close': 'exchange_rate',
        })

        wacc_df['report_date'] = pd.to_datetime(wacc_df['report_date']).astype('datetime64[us]')
        currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

        wacc_df = pd.merge_asof(
            wacc_df,
            currency_df,
            left_on='report_date',
            right_on='report_date',
            direction='backward'
        )
        wacc_df['total_debt_usd'] = round(wacc_df['total_debt'] / wacc_df['exchange_rate'], 0)
        wacc_df['interest_expense_usd'] = round(wacc_df['interest_expense'] / wacc_df['exchange_rate'], 0)
        wacc_df['pretax_income_usd'] = round(wacc_df['pretax_income'] / wacc_df['exchange_rate'], 0)
        wacc_df['tax_provision_usd'] = round(wacc_df['tax_provision'] / wacc_df['exchange_rate'], 0)

        market_cap_df = self.market_capitalization().drop(columns=['symbol'])

        market_cap_df['report_date'] = pd.to_datetime(market_cap_df['report_date']).astype('datetime64[us]')

        result_df1 = pd.merge_asof(
            wacc_df,
            market_cap_df,
            left_on='report_date',
            right_on='report_date',
            direction='backward'
        )

        max_date = wacc_df['report_date'].max()

        market_cap_after = market_cap_df.loc[
            (market_cap_df['report_date'] >= pd.Timestamp.today() - pd.DateOffset(years=5)) &
            (market_cap_df['report_date'] >= max_date)
        ]

        result_df2 = pd.merge_asof(
            market_cap_after,
            wacc_df,
            left_on='report_date',
            right_on='report_date',
            direction='backward'
        )
        result_df = pd.concat([result_df1, result_df2], ignore_index=True).drop_duplicates().sort_values('report_date').reset_index(drop=True)

        result_df = result_df[[
            'symbol',
            'report_date',
            'market_capitalization',
            'exchange_rate',
            'total_debt',
            'total_debt_usd',
            'interest_expense',
            'interest_expense_usd',
            'pretax_income',
            'pretax_income_usd',
            'tax_provision',
            'tax_provision_usd',
            'tax_rate_for_calcs'
        ]]
        ten_year_returns = sp500_cagr_returns_rolling(10)
        ten_year_returns['end_date'] = pd.to_datetime(ten_year_returns['end_date']).astype('datetime64[us]')

        result_df = pd.merge_asof(
            result_df,
            ten_year_returns,
            left_on='report_date',
            right_on='end_date',
            direction='backward'
        )

        result_df = result_df[[
            'symbol',
            'report_date',
            'market_capitalization',
            'exchange_rate',
            'total_debt',
            'total_debt_usd',
            'interest_expense',
            'interest_expense_usd',
            'pretax_income',
            'pretax_income_usd',
            'tax_provision',
            'tax_provision_usd',
            'tax_rate_for_calcs',
            'end_year',
            'cagr_returns_10_years'
        ]]

        result_df = result_df.rename(columns={
            'cagr_returns_10_years': 'sp500_10y_cagr',
            'end_year': 'sp500_cagr_end'
        })

        treasure = self.treasure.daily_treasure_yield()
        treasure['report_date'] = pd.to_datetime(treasure['report_date']).astype('datetime64[us]')

        result_df = pd.merge_asof(
            result_df,
            treasure,
            left_on='report_date',
            right_on='report_date',
            direction='backward'
        )

        result_df = result_df[[
            'symbol',
            'report_date',
            'market_capitalization',
            'exchange_rate',
            'total_debt',
            'total_debt_usd',
            'interest_expense',
            'interest_expense_usd',
            'pretax_income',
            'pretax_income_usd',
            'tax_provision',
            'tax_provision_usd',
            'tax_rate_for_calcs',
            'sp500_cagr_end',
            'sp500_10y_cagr',
            'bc10_year'
        ]]

        result_df = result_df.rename(columns={
            'bc10_year': 'treasure_10y_yield',
        })

        # Calculate 5-year beta using monthly returns
        result_df['beta_5y'] = self.beta("5y").iloc[0]['beta']

        result_df['tax_rate_for_calcs'] = np.where(
            result_df['tax_rate_for_calcs'].notna(),
            result_df['tax_rate_for_calcs'],
            result_df['tax_provision_usd'] / result_df['pretax_income_usd']
        )

        result_df['weight_of_debt'] = round(result_df['total_debt_usd'] / (result_df['total_debt_usd'] + result_df['market_capitalization']), 4)
        result_df['weight_of_equity'] = round(result_df['market_capitalization'] / (result_df['total_debt_usd'] + result_df['market_capitalization']), 4)
        result_df['cost_of_debt'] = round(result_df['interest_expense_usd'] / result_df['total_debt_usd'], 4)
        result_df['cost_of_equity'] = round(result_df['treasure_10y_yield'] + result_df['beta_5y'] * (result_df['sp500_10y_cagr'] - result_df['treasure_10y_yield']), 4)
        result_df['wacc'] = round(
            result_df['weight_of_debt'] * result_df['cost_of_debt'] * (1 - result_df['tax_rate_for_calcs']) +
            result_df['weight_of_equity'] * result_df['cost_of_equity'],
            4
        )
        return result_df

    def _add_discount_rate_section(self, ws, last_wacc, add_cell, add_border, bold, orange_fill, thin):
        """Create Discount Rate Estimates Section (rows 1-9).

        Populates the discount rate estimates including market cap, beta, debt,
        interest expense, pre-tax income, tax provision, risk-free rate, and
        expected market return. Also calculates WACC components.

        Args:
            ws: The openpyxl worksheet object.
            last_wacc: Dictionary containing the latest WACC data.
            add_cell: Helper function to add cell values with formatting.
            add_border: Helper function to add borders to cells.
            bold: Bold font style.
            orange_fill: Orange fill pattern for highlighted cells.
            thin: Border side style.
        """
        report_date = pd.to_datetime(last_wacc["report_date"]).strftime("%Y-%m-%d")
        row = 0
        add_cell("B", (row := row + 1), f"Discount Rate Estimates ({report_date})", font=bold)
        add_cell("B", (row := row + 1), "Market Cap (USD)", font=bold)
        add_cell("C", row, last_wacc['market_capitalization'], number_format='#,##0')
        add_cell("B", (row := row + 1), "β(5y)", font=bold)
        add_cell("C", row, last_wacc['beta_5y'], number_format='0.00')
        add_cell("B", (row := row + 1), "Total Debt (USD)", font=bold)
        add_cell("C", row, last_wacc['total_debt_usd'], number_format='#,##0')
        add_cell("B", (row := row + 1), "Interest Expense (USD)", font=bold)
        add_cell("C", row, last_wacc['interest_expense_usd'], number_format='#,##0')
        add_cell("B", (row := row + 1), "Pre-Tax Income (USD)", font=bold)
        add_cell("C", row, last_wacc['pretax_income_usd'], number_format='#,##0')
        add_cell("B", (row := row + 1), "Tax Provision", font=bold)
        add_cell("C", row, last_wacc['tax_provision_usd'], number_format='#,##0')
        add_cell("B", (row := row + 1), "Risk-Free Rate of Return (10Y Treasury Rate)", font=bold, fill=orange_fill)
        add_cell("C", row, last_wacc['treasure_10y_yield'], number_format='0.00%')
        add_cell("B", (row := row + 1), "Expected Market Return (S&P500 Avg Return)", font=bold)
        add_cell("C", row, last_wacc['sp500_10y_cagr'], number_format='0.00%')

        row = 1
        add_cell("D", (row := row + 1), "Weight of Debt", font=bold)
        add_cell("E", row, "=C4/(C2+C4)", number_format='0.00%')
        add_cell("D", (row := row + 1), "Weight of Equity", font=bold)
        add_cell("E", row, "=C2/(C2+C4)", number_format='0.00%')
        add_cell("D", (row := row + 1), "Cost of Debt", font=bold)
        add_cell("E", row, "=C5/C4", number_format='0.00%')
        add_cell("D", (row := row + 1), "Cost of Equity", font=bold)
        add_cell("E", row, "=C8+C3*(C9-C8)", number_format='0.00%')
        add_cell("D", (row := row + 1), "Tax Rate", font=bold)
        add_cell("E", row, last_wacc['tax_rate_for_calcs'], number_format='0.00%')
        add_cell("D", (row := row + 3), "WACC", font=bold, fill=orange_fill)
        add_cell("E", row, "=E2*E4*(1-E6)+E3*E5", number_format='0.00%')

        add_border(2, 9, ['B', 'C', 'D', 'E'])

    def _add_growth_estimates_section(self, ws, revenue_details, fcf_details, ebitda_details,
                                       net_income_details, finance_currency, add_cell, add_border,
                                       bold, orange_fill, thin) -> Dict[str, int]:
        """Create Growth Estimates Section with revenue, FCF, EBITDA, and net income growth data.

        Populates the growth estimates section showing 3-year historical data for each metric
        along with YoY growth rates and 3-year CAGR calculations.

        Args:
            ws: The openpyxl worksheet object.
            revenue_details: List of dicts with revenue data (date, value, yoy).
            fcf_details: List of dicts with FCF data (date, value, yoy).
            ebitda_details: List of dicts with EBITDA data (date, value, yoy).
            net_income_details: List of dicts with net income data (date, value, yoy).
            finance_currency: Currency code for display labels.
            add_cell: Helper function to add cell values with formatting.
            add_border: Helper function to add borders to cells.
            bold: Bold font style.
            orange_fill: Orange fill pattern for highlighted cells.
            thin: Border side style.

        Returns:
            Dict containing row numbers for CAGR calculations:
            - 'revenue_cagr_row': Row number of Revenue 3Y CAGR
            - 'fcf_cagr_row': Row number of FCF 3Y CAGR
            - 'ebitda_cagr_row': Row number of EBITDA 3Y CAGR
            - 'ni_cagr_row': Row number of Net Income 3Y CAGR
        """
        row = 0
        add_cell("G", (row := row + 1), f"Growth Estimates", font=bold)

        add_cell("G", (row := row + 1), f"Revenue ({finance_currency})", font=bold)
        y1_row = row + 1
        add_cell("G", (row := row + 1), revenue_details[0]['date'])
        add_cell("H", row, revenue_details[0]['value'], number_format='#,##0')
        add_cell("I", row, revenue_details[0]['yoy'], number_format='0.00%')
        add_cell("G", (row := row + 1), revenue_details[1]['date'])
        add_cell("H", row, revenue_details[1]['value'], number_format='#,##0')
        add_cell("I", row, revenue_details[1]['yoy'], number_format='0.00%')
        y3_row = row + 1
        add_cell("G", (row := row + 1), revenue_details[2]['date'])
        add_cell("H", row, revenue_details[2]['value'], number_format='#,##0')
        add_cell("I", row, revenue_details[2]['yoy'], number_format='0.00%')
        revenue_cagr_row = row + 1
        add_cell("G", (row := row + 1), "Revenue 3Y CAGR", font=bold)
        add_cell("H", row, f"=IF(H{y1_row}<=0,IF(H{y3_row}>0,\"Turned Positive\",\"N/A\"),IF(H{y3_row}<=0,\"Turned Negative\",POWER(H{y3_row}/H{y1_row},1/2)-1))", number_format='0.00%')

        row += 1
        add_cell("G", (row := row + 1), f"FCF ({finance_currency})", font=bold)
        fcf_y1_row = row + 1
        add_cell("G", (row := row + 1), fcf_details[0]['date'])
        add_cell("H", row, fcf_details[0]['value'], number_format='#,##0')
        add_cell("I", row, fcf_details[0]['yoy'], number_format='0.00%')
        add_cell("G", (row := row + 1), fcf_details[1]['date'])
        add_cell("H", row, fcf_details[1]['value'], number_format='#,##0')
        add_cell("I", row, fcf_details[1]['yoy'], number_format='0.00%')
        fcf_y3_row = row + 1
        add_cell("G", (row := row + 1), fcf_details[2]['date'])
        add_cell("H", row, fcf_details[2]['value'], number_format='#,##0')
        add_cell("I", row, fcf_details[2]['yoy'], number_format='0.00%')
        fcf_cagr_row = row + 1
        add_cell("G", (row := row + 1), "FCF 3Y CAGR", font=bold)
        add_cell("H", row, f"=IF(H{fcf_y1_row}<=0,IF(H{fcf_y3_row}>0,\"Turned Positive\",\"N/A\"),IF(H{fcf_y3_row}<=0,\"Turned Negative\",POWER(H{fcf_y3_row}/H{fcf_y1_row},1/2)-1))", number_format='0.00%')

        row += 1
        add_cell("G", (row := row + 1), f"EBITDA ({finance_currency})", font=bold)
        ebitda_y1_row = row + 1
        add_cell("G", (row := row + 1), ebitda_details[0]['date'])
        add_cell("H", row, ebitda_details[0]['value'], number_format='#,##0')
        add_cell("I", row, ebitda_details[0]['yoy'], number_format='0.00%')
        add_cell("G", (row := row + 1), ebitda_details[1]['date'])
        add_cell("H", row, ebitda_details[1]['value'], number_format='#,##0')
        add_cell("I", row, ebitda_details[1]['yoy'], number_format='0.00%')
        ebitda_y3_row = row + 1
        add_cell("G", (row := row + 1), ebitda_details[2]['date'])
        add_cell("H", row, ebitda_details[2]['value'], number_format='#,##0')
        add_cell("I", row, ebitda_details[2]['yoy'], number_format='0.00%')
        ebitda_cagr_row = row + 1
        add_cell("G", (row := row + 1), "EBITDA 3Y CAGR", font=bold)
        add_cell("H", row, f"=IF(H{ebitda_y1_row}<=0,IF(H{ebitda_y3_row}>0,\"Turned Positive\",\"N/A\"),IF(H{ebitda_y3_row}<=0,\"Turned Negative\",POWER(H{ebitda_y3_row}/H{ebitda_y1_row},1/2)-1))", number_format='0.00%')

        row += 1
        add_cell("G", (row := row + 1), f"Net Income ({finance_currency})", font=bold)
        ni_y1_row = row + 1
        add_cell("G", (row := row + 1), net_income_details[0]['date'])
        add_cell("H", row, net_income_details[0]['value'], number_format='#,##0')
        add_cell("I", row, net_income_details[0]['yoy'], number_format='0.00%')
        add_cell("G", (row := row + 1), net_income_details[1]['date'])
        add_cell("H", row, net_income_details[1]['value'], number_format='#,##0')
        add_cell("I", row, net_income_details[1]['yoy'], number_format='0.00%')
        ni_y3_row = row + 1
        add_cell("G", (row := row + 1), net_income_details[2]['date'])
        add_cell("H", row, net_income_details[2]['value'], number_format='#,##0')
        add_cell("I", row, net_income_details[2]['yoy'], number_format='0.00%')
        ni_cagr_row = row + 1
        add_cell("G", (row := row + 1), "Net Income 3Y CAGR", font=bold)
        add_cell("H", row, f"=IF(H{ni_y1_row}<=0,IF(H{ni_y3_row}>0,\"Turned Positive\",\"N/A\"),IF(H{ni_y3_row}<=0,\"Turned Negative\",POWER(H{ni_y3_row}/H{ni_y1_row},1/2)-1))", number_format='0.00%')

        add_border(2, row, ['G', 'H', 'I'])

        return {
            'revenue_cagr_row': revenue_cagr_row,
            'fcf_cagr_row': fcf_cagr_row,
            'ebitda_cagr_row': ebitda_cagr_row,
            'ni_cagr_row': ni_cagr_row
        }

    def _add_dcf_template_section(self, ws, base_fcf, end_date, revenue_cagr_row, fcf_cagr_row,
                                   ebitda_cagr_row, ni_cagr_row, revenue_details, add_cell,
                                   add_border, bold, orange_fill, thin) -> Dict[str, int]:
        """Create DCF Template Section with projections and historical FCF margins.

        Populates the DCF template including growth rate parameters, TTM revenue/FCF,
        projected FCF for years 1-10, terminal value calculation, and historical FCF margins.

        Args:
            ws: The openpyxl worksheet object.
            base_fcf: TTM free cash flow value.
            end_date: End date string for TTM period.
            revenue_cagr_row: Row number of Revenue 3Y CAGR.
            fcf_cagr_row: Row number of FCF 3Y CAGR.
            ebitda_cagr_row: Row number of EBITDA 3Y CAGR.
            ni_cagr_row: Row number of Net Income 3Y CAGR.
            revenue_details: List of dicts with revenue data for TTM calculations.
            add_cell: Helper function to add cell values with formatting.
            add_border: Helper function to add borders to cells.
            bold: Bold font style.
            orange_fill: Orange fill pattern for highlighted cells.
            thin: Border side style.

        Returns:
            Dict containing key row numbers:
            - 'total_value_row': Row number of Total Value
            - 'fcf_margin_row': Row number of FCF Margin
            - 'ttm_revenue_row': Row number of TTM Revenue
            - 'revenue_growth_1_5y_row': Row number of Future Revenue Growth (1-5Y)
            - 'revenue_growth_6_10y_row': Row number of Future Revenue Growth (6-10Y)
        """
        row = 15
        add_cell("B", (row := row + 1), "DCF Template", font=bold)

        decay_factor_row = row + 1
        add_cell("B", (row := row + 1), "Decay Factor (6~10Y)", font=bold, fill=orange_fill)
        add_cell("C", row, 0.9, number_format='0.00')

        growth_1_5y_row = row + 1
        add_cell("B", (row := row + 1), "Future Growth Rate (1~5 Years)", font=bold, fill=orange_fill)
        add_cell("C", row, f"=IFERROR((IF(ISNUMBER(H{revenue_cagr_row}),H{revenue_cagr_row},0)*0.4+IF(ISNUMBER(H{fcf_cagr_row}),H{fcf_cagr_row},0)*0.3+IF(ISNUMBER(H{ebitda_cagr_row}),H{ebitda_cagr_row},0)*0.2+IF(ISNUMBER(H{ni_cagr_row}),H{ni_cagr_row},0)*0.1)/(IF(ISNUMBER(H{revenue_cagr_row}),0.4,0)+IF(ISNUMBER(H{fcf_cagr_row}),0.3,0)+IF(ISNUMBER(H{ebitda_cagr_row}),0.2,0)+IF(ISNUMBER(H{ni_cagr_row}),0.1,0)),\"N/A\")",
                 number_format='0.00%')

        add_cell("B", (row := row + 1), "Future Growth Rate (6~10 Years)", font=bold, fill=orange_fill)
        add_cell("C", row, f"=MAX(C{growth_1_5y_row}*POWER(C{decay_factor_row},5),C8)", number_format='0.00%')

        add_cell("B", (row := row + 1), "Future Growth Rate (Terminal Stage)", font=bold, fill=orange_fill)
        add_cell("C", row, "=C8", number_format='0.00%')

        row = row + 1
        cell = ws[f"B{row}"]
        # Create rich text with normal and italic parts
        normal_text = TextBlock(InlineFont(b=True), "Discount Rate (%) (Default: WACC)\n")
        italic_text = TextBlock(InlineFont(b=True, i=True), "or S&P 500 Average Return")
        cell.value = CellRichText(normal_text, italic_text)
        cell.fill = orange_fill
        cell.alignment = Alignment(wrap_text=True)

        add_cell("C", row, "=E9", number_format='0.00%')

        ttm_revenue_df = self.ttm_revenue()
        if not ttm_revenue_df.empty:
            latest_ttm = ttm_revenue_df.iloc[-1]
            ttm_revenue_value = latest_ttm['ttm_total_revenue_usd']
            import json
            quarters = json.loads(latest_ttm['report_date_2_revenue'])
            quarter_dates = sorted(quarters.keys())
            start_date = pd.to_datetime(quarter_dates[0]).strftime("%Y-%m-%d")
            end_date = pd.to_datetime(quarter_dates[-1]).strftime("%Y-%m-%d")
            ttm_revenue_label = f"TTM Revenue (USD | {start_date} ~ {end_date})"
        else:
            ttm_revenue_value = 0
            ttm_revenue_label = "TTM Revenue N/A"

        ttm_revenue_row = row + 1
        add_cell("B", (row := row + 1), ttm_revenue_label, font=bold, fill=orange_fill)
        add_cell("C", row, ttm_revenue_value, number_format='#,##0')

        revenue_growth_1_5y_row = row + 1
        add_cell("B", (row := row + 1), "Future Revenue Growth Rate (1~5 Years)", font=bold, fill=orange_fill)
        add_cell("C", row, f"=H{revenue_cagr_row}", number_format='0.00%')

        revenue_growth_6_10y_row = row + 1
        add_cell("B", (row := row + 1), "Future Revenue Growth Rate (6~10 Years)", font=bold, fill=orange_fill)
        add_cell("C", row, f"=MAX(H{revenue_cagr_row}*POWER(C{decay_factor_row},5),C8)", number_format='0.00%')

        row += 1
        add_cell("B", (row := row + 1), "Year", font=bold)

        ttm_end_date = pd.to_datetime(end_date)
        right_align = Alignment(horizontal='right')
        add_cell("C", row, f"{ttm_end_date.strftime('%Y-%m-%d')} (TTM)", font=bold, alignment=right_align)

        for i, col in enumerate(['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M'], start=1):
            future_date = ttm_end_date + pd.DateOffset(years=i)
            add_cell(col, row, f"{future_date.year}/{future_date.month}/{future_date.day}", font=bold, alignment=right_align)

        fcf_row = row + 1
        add_cell("B", (row := row + 1), "Free Cash Flow (USD)", font=bold)

        add_cell("C", row, base_fcf, number_format='#,##0')

        growth_1_5y = f"C{growth_1_5y_row}"
        growth_6_10y = f"C{growth_1_5y_row + 1}"

        add_cell("D", row, f"=C{row}*(1+{growth_1_5y})", number_format='#,##0')
        add_cell("E", row, f"=D{row}*(1+{growth_1_5y})", number_format='#,##0')
        add_cell("F", row, f"=E{row}*(1+{growth_1_5y})", number_format='#,##0')
        add_cell("G", row, f"=F{row}*(1+{growth_1_5y})", number_format='#,##0')
        add_cell("H", row, f"=G{row}*(1+{growth_1_5y})", number_format='#,##0')
        add_cell("I", row, f"=H{row}*(1+{growth_6_10y})", number_format='#,##0')
        add_cell("J", row, f"=I{row}*(1+{growth_6_10y})", number_format='#,##0')
        add_cell("K", row, f"=J{row}*(1+{growth_6_10y})", number_format='#,##0')
        add_cell("L", row, f"=K{row}*(1+{growth_6_10y})", number_format='#,##0')
        add_cell("M", row, f"=L{row}*(1+{growth_6_10y})", number_format='#,##0')

        tv_row = row + 1
        add_cell("B", (row := row + 1), "Terminal Value (USD)", font=bold)

        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            add_cell(col, row, 0, number_format='#,##0')
        add_cell("M", row, f"=M{fcf_row}*(1 + C20) / (C9 - C20)", number_format='#,##0')

        total_value_row = row + 1
        add_cell("B", (row := row + 1), "Total Value (USD)", font=bold)

        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            add_cell(col, row, f"={col}{fcf_row}", number_format='#,##0')
        add_cell("M", row, f"=M{fcf_row}+M{tv_row}", number_format='#,##0')

        fcf_margin_row = row + 1
        add_cell("B", (row := row + 1), "FCF Margin", font=bold)

        revenue_growth_1_5y = f"C{revenue_growth_1_5y_row}"
        revenue_growth_6_10y = f"C{revenue_growth_6_10y_row}"

        add_cell("C", row, f"=C{fcf_row}/C{ttm_revenue_row}", number_format='0.00%')

        for i, col in enumerate(['D', 'E', 'F', 'G', 'H'], start=1):
            add_cell(col, row, f"={col}{fcf_row}/(C{ttm_revenue_row}*POWER(1+{revenue_growth_1_5y},{i}))", number_format='0.00%')

        for i, col in enumerate(['I', 'J', 'K', 'L', 'M'], start=6):
            add_cell(col, row, f"={col}{fcf_row}/(C{ttm_revenue_row}*POWER(1+{revenue_growth_1_5y},5)*POWER(1+{revenue_growth_6_10y},{i-5}))", number_format='0.00%')

        row += 1

        hist_col_count = 0
        fcf_margin_df = self.annual_fcf_margin()
        if not fcf_margin_df.empty:
            recent_fcf_margin = fcf_margin_df.tail(5).dropna(subset=['fcf_margin'])
            if not recent_fcf_margin.empty:
                year_history_row = row + 1
                add_cell("B", (row := row + 1), "Year (Historical)", font=bold)
                for i, (_, row_data) in enumerate(recent_fcf_margin.iterrows()):
                    if i >= 10:
                        break
                    col = chr(ord('C') + i)
                    year = pd.to_datetime(row_data['report_date']).strftime("%Y/%m/%d")
                    add_cell(col, row, year, font=bold, alignment=right_align)
                    hist_col_count = i + 1

                add_cell("B", (row := row + 1), "FCF Margin (Historical)", font=bold)
                for i, (_, row_data) in enumerate(recent_fcf_margin.iterrows()):
                    if i >= 10:
                        break
                    col = chr(ord('C') + i)
                    add_cell(col, row, row_data['fcf_margin'], number_format='0.00%')
            else:
                add_cell("B", (row := row + 1), "Year", font=bold)
                add_cell("B", (row := row + 1), "Historical FCF Margin", font=bold)
        else:
            add_cell("B", (row := row + 1), "Year", font=bold)
            add_cell("B", (row := row + 1), "Historical FCF Margin", font=bold)

        # Add border to DCF Template Section (irregular shape)
        dcf_start_row = 17
        params_end_row = revenue_growth_6_10y_row + 1
        year_row = params_end_row + 1
        proj_fcf_margin_row = year_row + 4
        history_end_row = row

        # Determine historical data columns
        hist_last_col = chr(ord('C') + hist_col_count - 1) if hist_col_count > 0 else 'B'

        # Top border: B17, C17
        ws[f'B{dcf_start_row}'].border = Border(top=thin, left=thin)
        ws[f'C{dcf_start_row}'].border = Border(top=thin, right=thin)

        # Left border: B18-B33
        for r in range(dcf_start_row + 1, history_end_row):
            ws[f'B{r}'].border = Border(left=thin)

        # Right border: C18-C25
        for r in range(dcf_start_row + 1, params_end_row + 1):
            cell = ws[f'C{r}']
            cell.border = Border(right=thin, top=cell.border.top, left=cell.border.left)

        # Top border of year row: D26-M26
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            cell = ws[f'{col}{year_row}']
            cell.border = Border(top=thin, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)

        # Right border: M26-M30
        for r in range(year_row, proj_fcf_margin_row + 1):
            cell = ws[f'M{r}']
            cell.border = Border(right=thin, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

        # Bottom border of FCF margin row: hist_last_col-M30
        for col_ord in range(ord(hist_last_col) + 1, ord('M') + 1):
            col = chr(col_ord)
            cell = ws[f'{col}{proj_fcf_margin_row}']
            cell.border = Border(bottom=thin, left=cell.border.left, right=cell.border.right, top=cell.border.top)

        # Historical section borders (if exists)
        if hist_col_count > 0:
            cell = ws[f'{hist_last_col}{proj_fcf_margin_row+1}']
            cell.border = Border(right=thin)
            history_start_row = proj_fcf_margin_row + 2

            # Right border: hist_last_col
            for r in range(history_start_row, history_end_row + 1):
                cell = ws[f'{hist_last_col}{r}']
                cell.border = Border(right=thin, top=cell.border.top, bottom=cell.border.bottom, left=cell.border.left)

            # Bottom border
            for i in range(ord('B'), ord(hist_last_col) + 1):
                col = chr(i)
                cell = ws[f'{col}{history_end_row}']
                cell.border = Border(bottom=thin, left=cell.border.left, right=cell.border.right, top=cell.border.top)
        else:
            # No historical data, add left and bottom border to the last row
            ws[f'B{history_end_row}'].border = Border(left=thin, bottom=thin)

        return {
            'total_value_row': total_value_row,
            'fcf_margin_row': fcf_margin_row,
            'ttm_revenue_row': ttm_revenue_row,
            'revenue_growth_1_5y_row': revenue_growth_1_5y_row,
            'revenue_growth_6_10y_row': revenue_growth_6_10y_row
        }

    def _add_dcf_value_section(self, ws, total_value_row, last_wacc, finance_currency,
                                add_cell, add_border, bold, orange_fill, thin) -> Dict[str, int]:
        """Create DCF Value Section with enterprise value, equity value, and fair price calculations.

        Populates the DCF value section including enterprise value (NPV of projected cash flows),
        cash and short-term investments, total debt, equity value, outstanding shares,
        fair price, current price, and margin of safety.

        Args:
            ws: The openpyxl worksheet object.
            total_value_row: Row number of Total Value in DCF template.
            last_wacc: Dictionary containing the latest WACC data.
            finance_currency: Currency code for balance sheet conversion.
            add_cell: Helper function to add cell values with formatting.
            add_border: Helper function to add borders to cells.
            bold: Bold font style.
            orange_fill: Orange fill pattern for highlighted cells.
            thin: Border side style.

        Returns:
            Dict containing key row numbers:
            - 'ev_row': Row number of Enterprise Value
            - 'cash_row': Row number of Cash & ST Investments
            - 'debt_row': Row number of Total Debt
            - 'equity_row': Row number of Equity Value
            - 'shares_row': Row number of Outstanding Shares
            - 'fair_price_row': Row number of Fair Price
            - 'current_price_row': Row number of Current Price
            - 'margin_row': Row number of Margin of Safety
        """
        row = 35
        report_date = pd.to_datetime(last_wacc["report_date"]).strftime("%Y-%m-%d")
        add_cell("B", (row := row + 1), f"DCF Value ({report_date})", font=bold)

        ev_row = row + 1
        add_cell("B", (row := row + 1), "Enterprise Value (USD)", font=bold, fill=orange_fill)
        add_cell("C", row, f"=NPV(C21,D{total_value_row}:M{total_value_row})", number_format='#,##0')

        cash_row = row + 1
        add_cell("B", (row := row + 1), "Cash & ST Investments (USD)", font=bold, fill=orange_fill)

        # Get cash value and convert to USD
        bs_df = self.quarterly_balance_sheet().df()
        cash_value = 0
        if not bs_df.empty:
            cash_rows = bs_df[bs_df['Breakdown'].str.contains('Cash, Cash Equivalents & Short Term Investments', na=False)]
            if not cash_rows.empty:
                date_columns = [col for col in bs_df.columns if col != 'Breakdown']
                if date_columns:
                    cash_value_original = cash_rows.iloc[0][date_columns[0]]
                    if pd.isna(cash_value_original) or cash_value_original == '*':
                        cash_value = 0
                    else:
                        # Get exchange rate for currency conversion

                        if finance_currency == 'USD':
                            cash_value = cash_value_original
                        else:
                            # Get latest quarter date from balance sheet
                            latest_bs_date = pd.to_datetime(date_columns[0])

                            # Get exchange rate
                            currency_df = self.currency(finance_currency + '=X')
                            currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

                            # Find exchange rate closest to balance sheet date
                            exchange_rate_row = currency_df[currency_df['report_date'] <= latest_bs_date].tail(1)
                            if not exchange_rate_row.empty:
                                _exchange_rate = exchange_rate_row.iloc[0]['close']
                                cash_value = round(float(cash_value_original) / float(_exchange_rate), 2)
                            else:
                                cash_value = float(cash_value_original)  # Fallback to original if no exchange rate found

        add_cell("C", row, cash_value, number_format='#,##0')

        debt_row = row + 1
        add_cell("B", (row := row + 1), "Total Debt", font=bold, fill=orange_fill)
        add_cell("C", row, "=C4", number_format='#,##0')

        equity_row = row + 1
        add_cell("B", (row := row + 1), "Equity Value", font=bold, fill=orange_fill)
        add_cell("C", row, f"=C{ev_row}+C{cash_row}-C{debt_row}", number_format='#,##0')

        shares_row = row + 1
        add_cell("B", (row := row + 1), "Outstanding Shares", font=bold, fill=orange_fill)
        mc_df = self.market_capitalization()
        shares_value = 0
        if not mc_df.empty:
            shares_value = mc_df.iloc[-1]['shares_outstanding']
            if pd.isna(shares_value):
                shares_value = 0
        add_cell("C", row, shares_value, number_format='#,##0')

        fair_price_row = row + 1
        add_cell("B", (row := row + 1), "Fair Price", font=bold, fill=orange_fill)
        add_cell("C", row, f"=C{equity_row}/C{shares_row}", number_format='0.00')

        current_price_row = row + 1
        add_cell("B", (row := row + 1), "Current Price", font=bold, fill=orange_fill)
        price_df = self.price()
        if not price_df.empty:
            current_price = price_df.iloc[-1]['close']
        else:
            current_price = 0
        add_cell("C", row, current_price, number_format='0.00')

        margin_row = row + 1
        add_cell("B", (row := row + 1), "Margin of safety", font=bold, fill=orange_fill)
        add_cell("C", row, f"=(C{fair_price_row}-C{current_price_row})/C{fair_price_row}", number_format='0.00%')
        add_border(37, row, ['B', 'C'])
        add_border(44, 44, ['B', 'C'], Side(style='thick', color='FFDD5E56'))

        return {
            'ev_row': ev_row,
            'cash_row': cash_row,
            'debt_row': debt_row,
            'equity_row': equity_row,
            'shares_row': shares_row,
            'fair_price_row': fair_price_row,
            'current_price_row': current_price_row,
            'margin_row': margin_row
        }

    def _add_key_metrics_display(self, ws, ev_row, cash_row, equity_row, shares_row,
                                  fair_price_row, current_price_row, margin_row, add_border):
        """Create Key Metrics Display with merged cells for fair price, current price, and buy/sell signal.

        Creates a visual summary section with merged cells displaying the fair price,
        current price, and a buy/sell recommendation based on comparing fair vs current price.
        Also adds conditional formatting to color the buy/sell signal.

        Args:
            ws: The openpyxl worksheet object.
            ev_row: Row number of Enterprise Value.
            cash_row: Row number of Cash & ST Investments.
            equity_row: Row number of Equity Value.
            shares_row: Row number of Outstanding Shares.
            fair_price_row: Row number of Fair Price.
            current_price_row: Row number of Current Price.
            margin_row: Row number of Margin of Safety.
            add_border: Helper function to add borders to cells.
        """
        # Merge cells in column E for key metrics display
        # Fair Price (E37:E38)
        ws.merge_cells(f'E{ev_row}:E{cash_row}')
        cell = ws[f'E{ev_row}']
        cell.value = f"Fair Price"
        cell.font = Font(size=15, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.number_format = '0.00'
        add_border(37, 38, ['E'], Side(style='thick', color='FF51A39A'))
        # Fair Price (F37:F38)
        ws.merge_cells(f'F{ev_row}:F{cash_row}')
        cell = ws[f'F{ev_row}']
        cell.value = f"=C{fair_price_row}"
        cell.font = Font(size=15, bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = '0.00'
        add_border(37, 38, ['F'], Side(style='thick', color='FF51A39A'))

        # Current Price (E40:E41)
        ws.merge_cells(f'E{equity_row}:E{shares_row}')
        cell = ws[f'E{equity_row}']
        cell.value = f"Current Price"
        cell.font = Font(size=15, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.number_format = '0.00'
        add_border(40, 41, ['E'], Side(style='thick', color='FF51A39A'))
        # Current Price (F40:F41)
        ws.merge_cells(f'F{equity_row}:F{shares_row}')
        cell = ws[f'F{equity_row}']
        cell.value = f"=C{current_price_row}"
        cell.font = Font(size=15, bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = '0.00'
        add_border(40, 41, ['F'], Side(style='thick', color='FF51A39A'))

        # Buy/Sell (E43:E44)
        ws.merge_cells(f'E{current_price_row}:E{margin_row}')
        cell = ws[f'E{current_price_row}']
        cell.value = f'Buy / Sell'
        cell.font = Font(size=15, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        add_border(43, 44, ['E'], Side(style='thick', color='FF51A39A'))
        # Buy/Sell (F43:F44)
        ws.merge_cells(f'F{current_price_row}:F{margin_row}')
        cell = ws[f'F{current_price_row}']
        cell.value = f'=IF(C{fair_price_row}>C{current_price_row},"Buy","Sell")'
        cell.font = Font(size=15, bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        add_border(43, 44, ['F'], Side(style='thick', color='FF51A39A'))

        # Add conditional formatting for Buy/Sell font color
        # Green font for "Buy"
        green_font = Font(color='FF51A39A', size=15, bold=True)
        buy_rule = CellIsRule(operator='equal', formula=['"Buy"'], font=green_font)
        ws.conditional_formatting.add(f'F{current_price_row}:F{margin_row}', buy_rule)

        # Red font for "Sell"
        red_font = Font(color='FFDD5E56', size=15, bold=True)
        sell_rule = CellIsRule(operator='equal', formula=['"Sell"'], font=red_font)
        ws.conditional_formatting.add(f'F{current_price_row}:F{margin_row}', sell_rule)

    def dcf(self) -> Dict[str, str]:
        """Generate a Discounted Cash Flow (DCF) valuation Excel spreadsheet.

        Creates a comprehensive DCF analysis workbook containing:
        - Discount Rate Estimates section with WACC calculation
        - Growth Estimates section with historical CAGR for revenue, FCF, EBITDA, and net income
        - DCF Template section with 10-year cash flow projections
        - DCF Value section with enterprise value, equity value, and fair price
        - Key metrics display with buy/sell recommendation

        Returns:
            Dict[str, str]: Dictionary containing:
                - file_path (str): Path to the generated Excel workbook
                - description (str): Description of the DCF analysis file
        """
        import json

        # Initialize workbook and styles
        wb = Workbook()
        ws = wb.active
        ws.title = f"DCF Value of {self.ticker}"
        bold = Font(bold=True)
        orange_fill = PatternFill(start_color="FFE6DB74", end_color="FFE6DB74", fill_type="solid")
        thin = Side(style='medium', color='FFB1B9F9')

        # Get company info and finance currency
        company_info = self.company_meta.get_company_info(self.ticker)
        finance_currency = (
            company_info.get("financial_currency")
            if company_info
            else "USD"
        )

        # Set column widths
        for col, width in zip(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"],
                             [1, 45, 18, 18, 18, 18, 18, 18, 18, 18, 18, 18, 18]):
            ws.column_dimensions[col].width = width

        # Helper function to add cell with formatting
        def add_cell(_col, _row, value, font=None, fill=None, number_format=None, alignment=None):
            _cell = ws[f"{_col}{_row}"]
            _cell.value = value
            if font:
                _cell.font = font
            if fill:
                _cell.fill = fill
            if number_format:
                _cell.number_format = number_format
            if alignment:
                _cell.alignment = alignment

        # Helper function to add borders
        def add_border(start_row, end_row, cols, border_side=None):
            if border_side is None:
                border_side = thin
            for _row in range(start_row, end_row + 1):
                for c in cols:
                    _cell = ws[f"{c}{_row}"]
                    left = border_side if c == cols[0] else None
                    right = border_side if c == cols[-1] else None
                    top = border_side if _row == start_row else None
                    bottom = border_side if _row == end_row else None
                    _cell.border = Border(left=left, right=right, top=top, bottom=bottom)

        # Helper function to extract growth details from dataframe
        def get_growth_details(growth_df, years=3):
            if growth_df.empty:
                return []
            recent = growth_df.tail(years)
            details = []
            for _, row_data in recent.iterrows():
                date_str = pd.to_datetime(row_data['report_date']).strftime("%Y-%m-%d")
                metric_name = [col for col in row_data.index if col not in ['symbol', 'report_date', 'yoy_growth'] and not col.startswith('prev_year_')][0]
                current_val = row_data.get(metric_name, 0)
                yoy = row_data.get('yoy_growth', 0)
                details.append({'date': date_str, 'value': current_val, 'yoy': yoy})
            while len(details) < years:
                details.insert(0, {'date': 'N/A', 'value': 0, 'yoy': 0})
            return details[-years:]

        # ========== Fetch Required Data ==========
        wacc = self.wacc()
        if wacc.empty:
            raise ValueError(
                f"Cannot calculate DCF for {self.ticker}: WACC data is unavailable. "
                f"This typically means the ticker has no financial statements (e.g. ETFs, indices)."
            )
        last_wacc = wacc.iloc[-1]

        revenue_growth = self.annual_revenue_yoy_growth()
        fcf_growth = self.annual_fcf_yoy_growth()
        ebitda_growth = self.annual_ebitda_yoy_growth()
        net_income_growth = self.annual_net_income_yoy_growth()

        revenue_details = get_growth_details(revenue_growth, 3)
        fcf_details = get_growth_details(fcf_growth, 3)
        ebitda_details = get_growth_details(ebitda_growth, 3)
        net_income_details = get_growth_details(net_income_growth, 3)

        # Get TTM FCF for base FCF value
        ttm_fcf_df = self.ttm_fcf()
        if not ttm_fcf_df.empty:
            latest_ttm = ttm_fcf_df.iloc[-1]
            base_fcf = latest_ttm['ttm_free_cash_flow_usd']
        else:
            base_fcf = 0

        # Get end_date from TTM revenue for DCF template
        ttm_revenue_df = self.ttm_revenue()
        if not ttm_revenue_df.empty:
            latest_ttm_rev = ttm_revenue_df.iloc[-1]
            quarters = json.loads(latest_ttm_rev['report_date_2_revenue'])
            quarter_dates = sorted(quarters.keys())
            end_date = pd.to_datetime(quarter_dates[-1]).strftime("%Y-%m-%d")
        else:
            end_date = pd.Timestamp.now().strftime("%Y-%m-%d")

        # ========== Add Discount Rate Estimates Section ==========
        self._add_discount_rate_section(ws, last_wacc, add_cell, add_border, bold, orange_fill, thin)

        # ========== Add Growth Estimates Section ==========
        growth_rows = self._add_growth_estimates_section(
            ws, revenue_details, fcf_details, ebitda_details, net_income_details,
            finance_currency, add_cell, add_border, bold, orange_fill, thin
        )

        # ========== Add DCF Template Section ==========
        template_rows = self._add_dcf_template_section(
            ws, base_fcf, end_date,
            growth_rows['revenue_cagr_row'], growth_rows['fcf_cagr_row'],
            growth_rows['ebitda_cagr_row'], growth_rows['ni_cagr_row'],
            revenue_details, add_cell, add_border, bold, orange_fill, thin
        )

        # ========== Add DCF Value Section ==========
        value_rows = self._add_dcf_value_section(
            ws, template_rows['total_value_row'], last_wacc, finance_currency,
            add_cell, add_border, bold, orange_fill, thin
        )

        # ========== Add Key Metrics Display ==========
        self._add_key_metrics_display(
            ws, value_rows['ev_row'], value_rows['cash_row'],
            value_rows['equity_row'], value_rows['shares_row'],
            value_rows['fair_price_row'], value_rows['current_price_row'],
            value_rows['margin_row'], add_border
        )

        # Save and return file path
        if in_notebook():
            # In notebook, save to current working directory
            output = f"{self.ticker}.xlsx"
        else:
            # In normal Python, save to DCF directory
            output = f"{validate_dcf_directory()}/{self.ticker}.xlsx"

        wb.save(output)
        wb.close()

        # Display download link and Google Drive button in notebook environment
        if in_notebook():
            from IPython.display import HTML, display
            import os

            # Get the filename for the URL
            filename = os.path.basename(output)

            # For notebook, we need to construct the URL to access the file
            # Use JavaScript to get the full URL dynamically
            download_and_drive_buttons = f"""
            <script src="https://apis.google.com/js/platform.js" async defer></script>
            <div style="margin-top: 12px;">
                <a href="{filename}" download="{self.ticker}_DCF.xlsx"
                   style="font-size:16px; margin-right:20px; display:inline-block;">
                   ⬇️ Download {self.ticker} DCF.xlsx
                </a>
                <div id="drive-button-{self.ticker}" style="display:inline-block;">
                    <div class="g-savetodrive"
                         data-src=""
                         data-filename="{self.ticker}_DCF.xlsx"
                         data-sitename="DefeatBeta DCF Analysis">
                    </div>
                </div>
            </div>
            <script>
            (function() {{
                // Get the current page URL and construct the full file URL
                var baseUrl = window.location.origin + window.location.pathname.replace(/\\/[^/]*$/, '');
                var fileUrl = baseUrl + '/files/{filename}';

                // Update the data-src attribute with the full URL
                var saveButton = document.querySelector('#drive-button-{self.ticker} .g-savetodrive');
                if (saveButton) {{
                    saveButton.setAttribute('data-src', fileUrl);

                    // Reload the Google Drive button to apply the new URL
                    if (typeof gapi !== 'undefined' && gapi.savetodrive) {{
                        gapi.savetodrive.go('#drive-button-{self.ticker}');
                    }}
                }}
            }})();
            </script>
            """
            display(HTML(download_and_drive_buttons))

        # Return file path and description
        return {
            'file_path': output,
            'description': f'DCF Valuation Analysis for {self.ticker}'
        }

    def industry_ttm_pe(self) -> pd.DataFrame:
        info = self.info()
        industry = info['industry']
        if isinstance(industry, pd.Series):
            industry = industry.iloc[0]

        if not industry or pd.isna(industry):
            raise ValueError(f"Unknown industry for this ticker: {self.ticker}")

        url = self.huggingface_client.get_url_path(stock_profile)
        sql = load_sql("select_tickers_by_industry", url=url, industry=industry)
        symbols = self.duckdb_client.query(sql)['symbol']
        symbols = symbols[symbols != self.ticker]
        symbols = pd.concat([pd.Series([self.ticker]), symbols], ignore_index=True)

        market_cap_table_sql = load_sql("select_market_cap_by_industry",
                                        stock_prices=self.huggingface_client.get_url_path(stock_prices),
                                        stock_shares_outstanding=self.huggingface_client.get_url_path(stock_shares_outstanding),
                                        symbols=", ".join(f"'{s}'" for s in symbols))

        # Keep wide format (report_date + one column per symbol) for paired exclusion
        market_cap_wide = self.duckdb_client.query(market_cap_table_sql)
        market_cap_wide = market_cap_wide.dropna(axis=1, how='all')
        market_cap_wide['report_date'] = pd.to_datetime(market_cap_wide['report_date']).astype('datetime64[us]')
        market_cap_cols = [col for col in market_cap_wide.columns if col != 'report_date']

        ttm_net_income_sql = load_sql("select_ttm_net_income_by_industry",
                                      stock_statement=self.huggingface_client.get_url_path(stock_statement),
                                      symbols=", ".join(f"'{s}'" for s in market_cap_cols))
        ttm_net_income_df = self.duckdb_client.query(ttm_net_income_sql).copy()


        currency_dict = self.company_meta.get_financial_currency_map()
        ttm_net_income_df['report_date'] = pd.to_datetime(ttm_net_income_df['report_date']).astype('datetime64[us]')

        usd_columns = []
        for symbol in ttm_net_income_df.columns:
            if symbol == 'report_date':
                continue
            currency = currency_dict.get(symbol, 'USD')
            if currency == 'USD':
                currency_df = pd.DataFrame({
                    'report_date': ttm_net_income_df['report_date'],
                    'close': 1.0
                })
            else:
                currency_df = self.currency(symbol=currency + '=X')
                currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

            merged_df = pd.merge_asof(
                ttm_net_income_df[['report_date', symbol]].rename(columns={symbol: 'ttm_net_income'}),
                currency_df,
                left_on='report_date',
                right_on='report_date',
                direction='backward'
            )
            usd_series = (merged_df['ttm_net_income'] / merged_df['close']).round(2)
            usd_series.name = f"{symbol}_usd"
            usd_columns.append(usd_series)

        # Keep TTM net income as wide format (report_date + one _usd column per symbol)
        ttm_net_income_usd_wide = pd.concat([ttm_net_income_df[['report_date']]] + usd_columns, axis=1)
        ttm_net_income_usd_wide = ttm_net_income_usd_wide.dropna(axis=1, how='all')
        # Sort by date BEFORE ffill, so values propagate forward in time correctly.
        # ffill after sort_values would apply in wrong direction if SQL returns rows
        # in non-chronological order.
        ttm_net_income_usd_wide = ttm_net_income_usd_wide.sort_values('report_date').reset_index(drop=True)
        ttm_net_income_usd_wide = ttm_net_income_usd_wide.ffill()

        # Sort market cap by date so row order matches ttm_aligned after merge_asof
        mc_reset = market_cap_wide.sort_values('report_date').reset_index(drop=True)

        # Align quarterly TTM data to daily market cap dates
        ttm_aligned = pd.merge_asof(
            mc_reset[['report_date']],
            ttm_net_income_usd_wide.sort_values('report_date'),
            on='report_date',
            direction='backward'
        ).reset_index(drop=True)

        # Paired exclusion: only include a company's market cap when its TTM net
        # income is also available, so numerator and denominator cover the same set
        # of companies for every date (MSCI methodology).
        paired_symbols = [s for s in market_cap_cols if f"{s}_usd" in ttm_aligned.columns]

        total_paired_market_cap = pd.Series(0.0, index=mc_reset.index)
        total_ttm_net_income = pd.Series(0.0, index=mc_reset.index)
        count_with_data = pd.Series(0, index=mc_reset.index)

        for s in paired_symbols:
            mc_col = mc_reset[s]
            ttm_col = ttm_aligned[f"{s}_usd"]
            # Exclude companies with negative TTM net income (standard aggregate
            # PE methodology: negative-earnings constituents are dropped so they
            # do not distort the industry-level ratio).
            has_data = ttm_col.notna() & mc_col.notna() & (ttm_col > 0)
            total_paired_market_cap += mc_col.where(has_data, 0.0)
            total_ttm_net_income += ttm_col.where(has_data, 0.0)
            count_with_data += has_data.astype(int)

        result = mc_reset[['report_date']].copy()
        result['industry'] = industry
        result['total_market_cap'] = total_paired_market_cap.where(count_with_data > 0)
        result['total_ttm_net_income'] = total_ttm_net_income.where(count_with_data > 0)
        result['industry_pe'] = (result['total_market_cap'] / result['total_ttm_net_income']).replace([np.inf, -np.inf], np.nan).round(2)
        # Drop rows where no company had sufficient data
        result = result.dropna(subset=['total_market_cap'])
        return result

    def industry_ps_ratio(self) -> pd.DataFrame:
        info = self.info()
        industry = info['industry']
        if isinstance(industry, pd.Series):
            industry = industry.iloc[0]

        if not industry or pd.isna(industry):
            raise ValueError(f"Unknown industry for this ticker: {self.ticker}")

        url = self.huggingface_client.get_url_path(stock_profile)
        sql = load_sql("select_tickers_by_industry", url=url, industry=industry)
        symbols = self.duckdb_client.query(sql)['symbol']
        symbols = symbols[symbols != self.ticker]
        symbols = pd.concat([pd.Series([self.ticker]), symbols], ignore_index=True)

        market_cap_table_sql = load_sql("select_market_cap_by_industry",
                                        stock_prices=self.huggingface_client.get_url_path(stock_prices),
                                        stock_shares_outstanding=self.huggingface_client.get_url_path(
                                            stock_shares_outstanding),
                                        symbols=", ".join(f"'{s}'" for s in symbols))

        # Keep wide format (report_date + one column per symbol) for paired exclusion
        market_cap_wide = self.duckdb_client.query(market_cap_table_sql)
        market_cap_wide = market_cap_wide.dropna(axis=1, how='all')
        market_cap_wide['report_date'] = pd.to_datetime(market_cap_wide['report_date']).astype('datetime64[us]')
        market_cap_cols = [col for col in market_cap_wide.columns if col != 'report_date']

        ttm_revenue_sql = load_sql("select_ttm_revenue_by_industry",
                                   stock_statement=self.huggingface_client.get_url_path(stock_statement),
                                   symbols=", ".join(f"'{s}'" for s in market_cap_cols))
        ttm_revenue_df = self.duckdb_client.query(ttm_revenue_sql).copy()

        currency_dict = self.company_meta.get_financial_currency_map()
        ttm_revenue_df['report_date'] = pd.to_datetime(ttm_revenue_df['report_date']).astype('datetime64[us]')

        usd_columns = []
        for symbol in ttm_revenue_df.columns:
            if symbol == 'report_date':
                continue
            currency = currency_dict.get(symbol, 'USD')
            if currency == 'USD':
                currency_df = pd.DataFrame({
                    'report_date': ttm_revenue_df['report_date'],
                    'close': 1.0
                })
            else:
                currency_df = self.currency(symbol=currency + '=X')
                currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

            merged_df = pd.merge_asof(
                ttm_revenue_df[['report_date', symbol]].rename(columns={symbol: 'ttm_revenue'}),
                currency_df,
                left_on='report_date',
                right_on='report_date',
                direction='backward'
            )
            usd_series = (merged_df['ttm_revenue'] / merged_df['close']).round(2)
            usd_series.name = f"{symbol}_usd"
            usd_columns.append(usd_series)

        # Keep TTM revenue as wide format (report_date + one _usd column per symbol)
        ttm_revenue_usd_wide = pd.concat([ttm_revenue_df[['report_date']]] + usd_columns, axis=1)
        ttm_revenue_usd_wide = ttm_revenue_usd_wide.dropna(axis=1, how='all')
        # Sort by date BEFORE ffill so values propagate forward in time correctly
        ttm_revenue_usd_wide = ttm_revenue_usd_wide.sort_values('report_date').reset_index(drop=True)
        ttm_revenue_usd_wide = ttm_revenue_usd_wide.ffill()

        # Sort market cap by date so row order matches ttm_aligned after merge_asof
        mc_reset = market_cap_wide.sort_values('report_date').reset_index(drop=True)

        # Align quarterly TTM data to daily market cap dates
        ttm_aligned = pd.merge_asof(
            mc_reset[['report_date']],
            ttm_revenue_usd_wide.sort_values('report_date'),
            on='report_date',
            direction='backward'
        ).reset_index(drop=True)

        # Paired exclusion: only include a company's market cap when its TTM revenue
        # is also available, so numerator and denominator cover the same set of
        # companies for every date (MSCI methodology).
        paired_symbols = [s for s in market_cap_cols if f"{s}_usd" in ttm_aligned.columns]

        total_paired_market_cap = pd.Series(0.0, index=mc_reset.index)
        total_ttm_revenue = pd.Series(0.0, index=mc_reset.index)
        count_with_data = pd.Series(0, index=mc_reset.index)

        for s in paired_symbols:
            mc_col = mc_reset[s]
            ttm_col = ttm_aligned[f"{s}_usd"]
            has_data = ttm_col.notna() & mc_col.notna()
            total_paired_market_cap += mc_col.where(has_data, 0.0)
            total_ttm_revenue += ttm_col.where(has_data, 0.0)
            count_with_data += has_data.astype(int)

        result = mc_reset[['report_date']].copy()
        result['industry'] = industry
        result['total_market_cap'] = total_paired_market_cap.where(count_with_data > 0)
        result['total_ttm_revenue'] = total_ttm_revenue.where(count_with_data > 0)
        result['industry_ps_ratio'] = (result['total_market_cap'] / result['total_ttm_revenue']).replace([np.inf, -np.inf], np.nan).round(2)
        # Drop rows where no company had sufficient data
        result = result.dropna(subset=['total_market_cap'])
        return result

    def industry_pb_ratio(self) -> pd.DataFrame:
        info = self.info()
        industry = info['industry']
        if isinstance(industry, pd.Series):
            industry = industry.iloc[0]

        if not industry or pd.isna(industry):
            raise ValueError(f"Unknown industry for this ticker: {self.ticker}")

        url = self.huggingface_client.get_url_path(stock_profile)
        sql = load_sql("select_tickers_by_industry", url=url, industry=industry)
        symbols = self.duckdb_client.query(sql)['symbol']
        symbols = symbols[symbols != self.ticker]
        symbols = pd.concat([pd.Series([self.ticker]), symbols], ignore_index=True)

        market_cap_table_sql = load_sql("select_market_cap_by_industry",
                                        stock_prices=self.huggingface_client.get_url_path(stock_prices),
                                        stock_shares_outstanding=self.huggingface_client.get_url_path(
                                            stock_shares_outstanding),
                                        symbols=", ".join(f"'{s}'" for s in symbols))

        # Keep wide format (report_date + one column per symbol) for paired exclusion
        market_cap_wide = self.duckdb_client.query(market_cap_table_sql)
        market_cap_wide = market_cap_wide.dropna(axis=1, how='all')
        market_cap_wide['report_date'] = pd.to_datetime(market_cap_wide['report_date']).astype('datetime64[us]')
        market_cap_cols = [col for col in market_cap_wide.columns if col != 'report_date']

        bve_sql = load_sql("select_quarterly_book_value_of_equity_by_industry",
                           stock_statement=self.huggingface_client.get_url_path(stock_statement),
                           symbols=", ".join(f"'{s}'" for s in market_cap_cols))
        bve_df = self.duckdb_client.query(bve_sql).copy()

        currency_dict = self.company_meta.get_financial_currency_map()
        bve_df['report_date'] = pd.to_datetime(bve_df['report_date']).astype('datetime64[us]')

        usd_columns = []
        for symbol in bve_df.columns:
            if symbol == 'report_date':
                continue
            currency = currency_dict.get(symbol, 'USD')
            if currency == 'USD':
                currency_df = pd.DataFrame({
                    'report_date': bve_df['report_date'],
                    'close': 1.0
                })
            else:
                currency_df = self.currency(symbol=currency + '=X')
                currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')

            merged_df = pd.merge_asof(
                bve_df[['report_date', symbol]].rename(columns={symbol: 'bve'}),
                currency_df,
                left_on='report_date',
                right_on='report_date',
                direction='backward'
            )
            usd_series = (merged_df['bve'] / merged_df['close']).round(2)
            usd_series.name = f"{symbol}_usd"
            usd_columns.append(usd_series)

        # Keep BVE as wide format (report_date + one _usd column per symbol)
        bve_usd_wide = pd.concat([bve_df[['report_date']]] + usd_columns, axis=1)
        bve_usd_wide = bve_usd_wide.dropna(axis=1, how='all')
        # Sort by date BEFORE ffill so values propagate forward in time correctly
        bve_usd_wide = bve_usd_wide.sort_values('report_date').reset_index(drop=True)
        bve_usd_wide = bve_usd_wide.ffill()

        # Sort market cap by date so row order matches bve_aligned after merge_asof
        mc_reset = market_cap_wide.sort_values('report_date').reset_index(drop=True)

        # Align quarterly BVE data to daily market cap dates
        bve_aligned = pd.merge_asof(
            mc_reset[['report_date']],
            bve_usd_wide.sort_values('report_date'),
            on='report_date',
            direction='backward'
        ).reset_index(drop=True)

        # Paired exclusion: only include a company's market cap when its BVE is also
        # available and positive. Negative book value distorts the industry-level ratio
        # (same rationale as excluding negative earnings for PE).
        paired_symbols = [s for s in market_cap_cols if f"{s}_usd" in bve_aligned.columns]

        total_paired_market_cap = pd.Series(0.0, index=mc_reset.index)
        total_bve = pd.Series(0.0, index=mc_reset.index)
        count_with_data = pd.Series(0, index=mc_reset.index)

        for s in paired_symbols:
            mc_col = mc_reset[s]
            bve_col = bve_aligned[f"{s}_usd"]
            has_data = bve_col.notna() & mc_col.notna() & (bve_col > 0)
            total_paired_market_cap += mc_col.where(has_data, 0.0)
            total_bve += bve_col.where(has_data, 0.0)
            count_with_data += has_data.astype(int)

        result = mc_reset[['report_date']].copy()
        result['industry'] = industry
        result['total_market_cap'] = total_paired_market_cap.where(count_with_data > 0)
        result['total_bve'] = total_bve.where(count_with_data > 0)
        result['industry_pb_ratio'] = (result['total_market_cap'] / result['total_bve']).replace([np.inf, -np.inf], np.nan).round(2)
        # Drop rows where no company had sufficient data
        result = result.dropna(subset=['total_market_cap'])
        return result

    def industry_roe(self) -> pd.DataFrame:
        info = self.info()
        industry = info['industry']
        if isinstance(industry, pd.Series):
            industry = industry.iloc[0]

        if not industry or pd.isna(industry):
            raise ValueError(f"Unknown industry for this ticker: {self.ticker}")

        url = self.huggingface_client.get_url_path(stock_profile)
        sql = load_sql("select_tickers_by_industry", url=url, industry=industry)
        symbols = self.duckdb_client.query(sql)['symbol']
        symbols = symbols[symbols != self.ticker]
        symbols = pd.concat([pd.Series([self.ticker]), symbols], ignore_index=True)

        ttm_roe_sql = load_sql("select_ttm_roe_by_industry",
                               stock_statement=self.huggingface_client.get_url_path(stock_statement),
                               symbols=", ".join(f"'{s}'" for s in symbols))
        ttm_roe_table = self.duckdb_client.query(ttm_roe_sql)

        ttm_roe_table['report_date'] = pd.to_datetime(ttm_roe_table['report_date']).astype('datetime64[us]')
        ttm_roe_table = ttm_roe_table.sort_values('report_date').reset_index(drop=True)

        currency_dict = self.company_meta.get_financial_currency_map()

        ni_suffix = '_ttm_net_income'
        eq_suffix = '_ttm_avg_equity'
        ni_symbols = [col[:-len(ni_suffix)] for col in ttm_roe_table.columns if col.endswith(ni_suffix)]

        # Monthly date baseline: every month end from first to last fiscal quarter in the data
        min_date = ttm_roe_table['report_date'].min()
        max_date = ttm_roe_table['report_date'].max()
        baseline = pd.DataFrame({
            'report_date': pd.date_range(start=min_date, end=max_date, freq='ME').astype('datetime64[us]')
        })

        total_net_income = np.zeros(len(baseline))
        total_avg_equity = np.zeros(len(baseline))
        count_with_data = np.zeros(len(baseline), dtype=int)

        for symbol in ni_symbols:
            ni_col = f"{symbol}{ni_suffix}"
            eq_col = f"{symbol}{eq_suffix}"
            if ni_col not in ttm_roe_table.columns or eq_col not in ttm_roe_table.columns:
                continue

            currency = currency_dict.get(symbol, 'USD')
            if currency == 'USD':
                fx_series = pd.Series(1.0, index=ttm_roe_table.index)
            else:
                currency_df = self.currency(symbol=currency + '=X')
                currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')
                fx_series = pd.merge_asof(
                    ttm_roe_table[['report_date']],
                    currency_df,
                    on='report_date',
                    direction='backward'
                )['close']

            symbol_df = pd.DataFrame({
                'report_date': ttm_roe_table['report_date'],
                'ni_usd': (ttm_roe_table[ni_col] / fx_series).values,
                'eq_usd': (ttm_roe_table[eq_col] / fx_series).values,
            }).dropna()

            # Align each company's quarterly TTM data to the monthly baseline via forward-fill
            ni_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'ni_usd']],
                on='report_date', direction='backward'
            )['ni_usd']
            eq_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'eq_usd']],
                on='report_date', direction='backward'
            )['eq_usd']

            # Paired exclusion: only include when both TTM net income and avg equity are available
            has_data = ni_aligned.notna().values & eq_aligned.notna().values
            total_net_income += np.where(has_data, ni_aligned.fillna(0).values, 0.0)
            total_avg_equity += np.where(has_data, eq_aligned.fillna(0).values, 0.0)
            count_with_data += has_data.astype(int)

        result = baseline.copy()
        result.insert(1, 'industry', industry)
        has_any = count_with_data > 0
        result['total_ttm_net_income'] = np.where(has_any, total_net_income, np.nan)
        result['total_avg_equity'] = np.where(has_any, total_avg_equity, np.nan)
        result['industry_roe'] = (result['total_ttm_net_income'] / result['total_avg_equity']).replace([np.inf, -np.inf], np.nan).round(4)
        result = result.dropna(subset=['total_ttm_net_income'])
        return result

    def industry_roa(self) -> pd.DataFrame:
        info = self.info()
        industry = info['industry']
        if isinstance(industry, pd.Series):
            industry = industry.iloc[0]

        if not industry or pd.isna(industry):
            raise ValueError(f"Unknown industry for this ticker: {self.ticker}")

        url = self.huggingface_client.get_url_path(stock_profile)
        sql = load_sql("select_tickers_by_industry", url=url, industry=industry)
        symbols = self.duckdb_client.query(sql)['symbol']
        symbols = symbols[symbols != self.ticker]
        symbols = pd.concat([pd.Series([self.ticker]), symbols], ignore_index=True)

        ttm_roa_sql = load_sql("select_ttm_roa_by_industry",
                               stock_statement=self.huggingface_client.get_url_path(stock_statement),
                               symbols=", ".join(f"'{s}'" for s in symbols))
        ttm_roa_table = self.duckdb_client.query(ttm_roa_sql)

        ttm_roa_table['report_date'] = pd.to_datetime(ttm_roa_table['report_date']).astype('datetime64[us]')
        ttm_roa_table = ttm_roa_table.sort_values('report_date').reset_index(drop=True)

        currency_dict = self.company_meta.get_financial_currency_map()

        ni_suffix = '_ttm_net_income'
        assets_suffix = '_ttm_avg_assets'
        ni_symbols = [col[:-len(ni_suffix)] for col in ttm_roa_table.columns if col.endswith(ni_suffix)]

        # Monthly date baseline: every month end from first to last fiscal quarter in the data
        min_date = ttm_roa_table['report_date'].min()
        max_date = ttm_roa_table['report_date'].max()
        baseline = pd.DataFrame({
            'report_date': pd.date_range(start=min_date, end=max_date, freq='ME').astype('datetime64[us]')
        })

        total_net_income = np.zeros(len(baseline))
        total_avg_assets = np.zeros(len(baseline))
        count_with_data = np.zeros(len(baseline), dtype=int)

        for symbol in ni_symbols:
            ni_col = f"{symbol}{ni_suffix}"
            assets_col = f"{symbol}{assets_suffix}"
            if ni_col not in ttm_roa_table.columns or assets_col not in ttm_roa_table.columns:
                continue

            currency = currency_dict.get(symbol, 'USD')
            if currency == 'USD':
                fx_series = pd.Series(1.0, index=ttm_roa_table.index)
            else:
                currency_df = self.currency(symbol=currency + '=X')
                currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')
                fx_series = pd.merge_asof(
                    ttm_roa_table[['report_date']],
                    currency_df,
                    on='report_date',
                    direction='backward'
                )['close']

            symbol_df = pd.DataFrame({
                'report_date': ttm_roa_table['report_date'],
                'ni_usd': (ttm_roa_table[ni_col] / fx_series).values,
                'assets_usd': (ttm_roa_table[assets_col] / fx_series).values,
            }).dropna()

            # Align each company's quarterly TTM data to the monthly baseline via forward-fill
            ni_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'ni_usd']],
                on='report_date', direction='backward'
            )['ni_usd']
            assets_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'assets_usd']],
                on='report_date', direction='backward'
            )['assets_usd']

            # Paired exclusion: only include when both TTM net income and avg assets are available
            has_data = ni_aligned.notna().values & assets_aligned.notna().values
            total_net_income += np.where(has_data, ni_aligned.fillna(0).values, 0.0)
            total_avg_assets += np.where(has_data, assets_aligned.fillna(0).values, 0.0)
            count_with_data += has_data.astype(int)

        result = baseline.copy()
        result.insert(1, 'industry', industry)
        has_any = count_with_data > 0
        result['total_ttm_net_income'] = np.where(has_any, total_net_income, np.nan)
        result['total_avg_assets'] = np.where(has_any, total_avg_assets, np.nan)
        result['industry_roa'] = (result['total_ttm_net_income'] / result['total_avg_assets']).replace([np.inf, -np.inf], np.nan).round(4)
        result = result.dropna(subset=['total_ttm_net_income'])
        return result

    def industry_roic(self) -> pd.DataFrame:
        info = self.info()
        industry = info['industry']
        if isinstance(industry, pd.Series):
            industry = industry.iloc[0]

        if not industry or pd.isna(industry):
            raise ValueError(f"Unknown industry for this ticker: {self.ticker}")

        url = self.huggingface_client.get_url_path(stock_profile)
        sql = load_sql("select_tickers_by_industry", url=url, industry=industry)
        symbols = self.duckdb_client.query(sql)['symbol']
        symbols = symbols[symbols != self.ticker]
        symbols = pd.concat([pd.Series([self.ticker]), symbols], ignore_index=True)

        ttm_roic_sql = load_sql("select_ttm_roic_by_industry",
                                stock_statement=self.huggingface_client.get_url_path(stock_statement),
                                symbols=", ".join(f"'{s}'" for s in symbols))
        ttm_roic_table = self.duckdb_client.query(ttm_roic_sql)

        ttm_roic_table['report_date'] = pd.to_datetime(ttm_roic_table['report_date']).astype('datetime64[us]')
        ttm_roic_table = ttm_roic_table.sort_values('report_date').reset_index(drop=True)

        currency_dict = self.company_meta.get_financial_currency_map()

        nopat_suffix = '_ttm_nopat'
        ic_suffix = '_ttm_avg_invested_capital'
        nopat_symbols = [col[:-len(nopat_suffix)] for col in ttm_roic_table.columns if col.endswith(nopat_suffix)]

        # Monthly date baseline: every month end from first to last fiscal quarter in the data
        min_date = ttm_roic_table['report_date'].min()
        max_date = ttm_roic_table['report_date'].max()
        baseline = pd.DataFrame({
            'report_date': pd.date_range(start=min_date, end=max_date, freq='ME').astype('datetime64[us]')
        })

        total_ttm_nopat = np.zeros(len(baseline))
        total_avg_invested_capital = np.zeros(len(baseline))
        count_with_data = np.zeros(len(baseline), dtype=int)

        for symbol in nopat_symbols:
            nopat_col = f"{symbol}{nopat_suffix}"
            ic_col = f"{symbol}{ic_suffix}"
            if nopat_col not in ttm_roic_table.columns or ic_col not in ttm_roic_table.columns:
                continue

            currency = currency_dict.get(symbol, 'USD')
            if currency == 'USD':
                fx_series = pd.Series(1.0, index=ttm_roic_table.index)
            else:
                currency_df = self.currency(symbol=currency + '=X')
                currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')
                fx_series = pd.merge_asof(
                    ttm_roic_table[['report_date']],
                    currency_df,
                    on='report_date',
                    direction='backward'
                )['close']

            symbol_df = pd.DataFrame({
                'report_date': ttm_roic_table['report_date'],
                'nopat_usd': (ttm_roic_table[nopat_col] / fx_series).values,
                'ic_usd': (ttm_roic_table[ic_col] / fx_series).values,
            }).dropna()

            # Align each company's quarterly TTM data to the monthly baseline via forward-fill
            nopat_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'nopat_usd']],
                on='report_date', direction='backward'
            )['nopat_usd']
            ic_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'ic_usd']],
                on='report_date', direction='backward'
            )['ic_usd']

            # Paired exclusion: only include when both TTM nopat and avg invested capital are available
            has_data = nopat_aligned.notna().values & ic_aligned.notna().values
            total_ttm_nopat += np.where(has_data, nopat_aligned.fillna(0).values, 0.0)
            total_avg_invested_capital += np.where(has_data, ic_aligned.fillna(0).values, 0.0)
            count_with_data += has_data.astype(int)

        result = baseline.copy()
        result.insert(1, 'industry', industry)
        has_any = count_with_data > 0
        result['total_ttm_nopat'] = np.where(has_any, total_ttm_nopat, np.nan)
        result['total_avg_invested_capital'] = np.where(has_any, total_avg_invested_capital, np.nan)
        result['industry_roic'] = (result['total_ttm_nopat'] / result['total_avg_invested_capital']).replace([np.inf, -np.inf], np.nan).round(4)
        result = result.dropna(subset=['total_ttm_nopat'])
        return result

    def industry_equity_multiplier(self) -> pd.DataFrame:
        roe = self.industry_roe()
        roa = self.industry_roa()

        roe['report_date'] = pd.to_datetime(roe['report_date']).astype('datetime64[us]')
        roa['report_date'] = pd.to_datetime(roa['report_date']).astype('datetime64[us]')

        # Both use the same monthly baseline, so a regular inner merge aligns perfectly
        result_df = roe[['report_date', 'industry', 'industry_roe']].merge(
            roa[['report_date', 'industry_roa']],
            on='report_date',
            how='inner'
        )

        result_df['industry_equity_multiplier'] = (result_df['industry_roe'] / result_df['industry_roa']).replace([np.inf, -np.inf], np.nan).round(2)
        return result_df

    def industry_quarterly_gross_margin(self) -> pd.DataFrame:
        info = self.info()
        industry = info['industry']
        if isinstance(industry, pd.Series):
            industry = industry.iloc[0]

        if not industry or pd.isna(industry):
            raise ValueError(f"Unknown industry for this ticker: {self.ticker}")

        url = self.huggingface_client.get_url_path(stock_profile)
        sql = load_sql("select_tickers_by_industry", url=url, industry=industry)
        symbols = self.duckdb_client.query(sql)['symbol']
        symbols = symbols[symbols != self.ticker]
        symbols = pd.concat([pd.Series([self.ticker]), symbols], ignore_index=True)

        ttm_sql = load_sql("select_ttm_gross_margin_by_industry",
                           stock_statement=self.huggingface_client.get_url_path(stock_statement),
                           symbols=", ".join(f"'{s}'" for s in symbols))
        ttm_table = self.duckdb_client.query(ttm_sql)

        ttm_table['report_date'] = pd.to_datetime(ttm_table['report_date']).astype('datetime64[us]')
        ttm_table = ttm_table.sort_values('report_date').reset_index(drop=True)

        currency_dict = self.company_meta.get_financial_currency_map()

        gp_suffix = '_ttm_gross_profit'
        rev_suffix = '_ttm_revenue'
        gp_symbols = [col[:-len(gp_suffix)] for col in ttm_table.columns if col.endswith(gp_suffix)]

        # Monthly date baseline: every month end from first to last fiscal quarter in the data
        min_date = ttm_table['report_date'].min()
        max_date = ttm_table['report_date'].max()
        baseline = pd.DataFrame({
            'report_date': pd.date_range(start=min_date, end=max_date, freq='ME').astype('datetime64[us]')
        })

        total_gross_profit = np.zeros(len(baseline))
        total_revenue = np.zeros(len(baseline))
        count_with_data = np.zeros(len(baseline), dtype=int)

        for symbol in gp_symbols:
            gp_col = f"{symbol}{gp_suffix}"
            rev_col = f"{symbol}{rev_suffix}"
            if gp_col not in ttm_table.columns or rev_col not in ttm_table.columns:
                continue

            currency = currency_dict.get(symbol, 'USD')
            if currency == 'USD':
                fx_series = pd.Series(1.0, index=ttm_table.index)
            else:
                currency_df = self.currency(symbol=currency + '=X')
                currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')
                fx_series = pd.merge_asof(
                    ttm_table[['report_date']],
                    currency_df,
                    on='report_date',
                    direction='backward'
                )['close']

            symbol_df = pd.DataFrame({
                'report_date': ttm_table['report_date'],
                'gp_usd': (ttm_table[gp_col] / fx_series).values,
                'rev_usd': (ttm_table[rev_col] / fx_series).values,
            }).dropna()

            # Align each company's quarterly TTM data to the monthly baseline via forward-fill
            gp_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'gp_usd']],
                on='report_date', direction='backward'
            )['gp_usd']
            rev_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'rev_usd']],
                on='report_date', direction='backward'
            )['rev_usd']

            # Paired exclusion: only include when both TTM gross profit and revenue are available
            has_data = gp_aligned.notna().values & rev_aligned.notna().values
            total_gross_profit += np.where(has_data, gp_aligned.fillna(0).values, 0.0)
            total_revenue += np.where(has_data, rev_aligned.fillna(0).values, 0.0)
            count_with_data += has_data.astype(int)

        result = baseline.copy()
        result.insert(1, 'industry', industry)
        has_any = count_with_data > 0
        result['total_ttm_gross_profit'] = np.where(has_any, total_gross_profit, np.nan)
        result['total_ttm_revenue'] = np.where(has_any, total_revenue, np.nan)
        result['industry_gross_margin'] = (result['total_ttm_gross_profit'] / result['total_ttm_revenue']).replace([np.inf, -np.inf], np.nan).round(4)
        result = result.dropna(subset=['total_ttm_gross_profit'])
        return result

    def industry_quarterly_ebitda_margin(self) -> pd.DataFrame:
        info = self.info()
        industry = info['industry']
        if isinstance(industry, pd.Series):
            industry = industry.iloc[0]

        if not industry or pd.isna(industry):
            raise ValueError(f"Unknown industry for this ticker: {self.ticker}")

        url = self.huggingface_client.get_url_path(stock_profile)
        sql = load_sql("select_tickers_by_industry", url=url, industry=industry)
        symbols = self.duckdb_client.query(sql)['symbol']
        symbols = symbols[symbols != self.ticker]
        symbols = pd.concat([pd.Series([self.ticker]), symbols], ignore_index=True)

        ttm_sql = load_sql("select_ttm_ebitda_margin_by_industry",
                           stock_statement=self.huggingface_client.get_url_path(stock_statement),
                           symbols=", ".join(f"'{s}'" for s in symbols))
        ttm_table = self.duckdb_client.query(ttm_sql)

        ttm_table['report_date'] = pd.to_datetime(ttm_table['report_date']).astype('datetime64[us]')
        ttm_table = ttm_table.sort_values('report_date').reset_index(drop=True)

        currency_dict = self.company_meta.get_financial_currency_map()

        ebitda_suffix = '_ttm_ebitda'
        rev_suffix = '_ttm_revenue'
        ebitda_symbols = [col[:-len(ebitda_suffix)] for col in ttm_table.columns if col.endswith(ebitda_suffix)]

        # Monthly date baseline: every month end from first to last fiscal quarter in the data
        min_date = ttm_table['report_date'].min()
        max_date = ttm_table['report_date'].max()
        baseline = pd.DataFrame({
            'report_date': pd.date_range(start=min_date, end=max_date, freq='ME').astype('datetime64[us]')
        })

        total_ebitda = np.zeros(len(baseline))
        total_revenue = np.zeros(len(baseline))
        count_with_data = np.zeros(len(baseline), dtype=int)

        for symbol in ebitda_symbols:
            ebitda_col = f"{symbol}{ebitda_suffix}"
            rev_col = f"{symbol}{rev_suffix}"
            if ebitda_col not in ttm_table.columns or rev_col not in ttm_table.columns:
                continue

            currency = currency_dict.get(symbol, 'USD')
            if currency == 'USD':
                fx_series = pd.Series(1.0, index=ttm_table.index)
            else:
                currency_df = self.currency(symbol=currency + '=X')
                currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')
                fx_series = pd.merge_asof(
                    ttm_table[['report_date']],
                    currency_df,
                    on='report_date',
                    direction='backward'
                )['close']

            symbol_df = pd.DataFrame({
                'report_date': ttm_table['report_date'],
                'ebitda_usd': (ttm_table[ebitda_col] / fx_series).values,
                'rev_usd': (ttm_table[rev_col] / fx_series).values,
            }).dropna()

            # Align each company's quarterly TTM data to the monthly baseline via forward-fill
            ebitda_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'ebitda_usd']],
                on='report_date', direction='backward'
            )['ebitda_usd']
            rev_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'rev_usd']],
                on='report_date', direction='backward'
            )['rev_usd']

            # Paired exclusion: only include when both TTM ebitda and revenue are available
            has_data = ebitda_aligned.notna().values & rev_aligned.notna().values
            total_ebitda += np.where(has_data, ebitda_aligned.fillna(0).values, 0.0)
            total_revenue += np.where(has_data, rev_aligned.fillna(0).values, 0.0)
            count_with_data += has_data.astype(int)

        result = baseline.copy()
        result.insert(1, 'industry', industry)
        has_any = count_with_data > 0
        result['total_ttm_ebitda'] = np.where(has_any, total_ebitda, np.nan)
        result['total_ttm_revenue'] = np.where(has_any, total_revenue, np.nan)
        result['industry_ebitda_margin'] = (result['total_ttm_ebitda'] / result['total_ttm_revenue']).replace([np.inf, -np.inf], np.nan).round(4)
        result = result.dropna(subset=['total_ttm_ebitda'])
        return result

    def industry_quarterly_net_margin(self) -> pd.DataFrame:
        info = self.info()
        industry = info['industry']
        if isinstance(industry, pd.Series):
            industry = industry.iloc[0]

        if not industry or pd.isna(industry):
            raise ValueError(f"Unknown industry for this ticker: {self.ticker}")

        url = self.huggingface_client.get_url_path(stock_profile)
        sql = load_sql("select_tickers_by_industry", url=url, industry=industry)
        symbols = self.duckdb_client.query(sql)['symbol']
        symbols = symbols[symbols != self.ticker]
        symbols = pd.concat([pd.Series([self.ticker]), symbols], ignore_index=True)

        ttm_sql = load_sql("select_ttm_net_margin_by_industry",
                           stock_statement=self.huggingface_client.get_url_path(stock_statement),
                           symbols=", ".join(f"'{s}'" for s in symbols))
        ttm_table = self.duckdb_client.query(ttm_sql)

        ttm_table['report_date'] = pd.to_datetime(ttm_table['report_date']).astype('datetime64[us]')
        ttm_table = ttm_table.sort_values('report_date').reset_index(drop=True)

        currency_dict = self.company_meta.get_financial_currency_map()

        ni_suffix = '_ttm_net_income'
        rev_suffix = '_ttm_revenue'
        ni_symbols = [col[:-len(ni_suffix)] for col in ttm_table.columns if col.endswith(ni_suffix)]

        # Monthly date baseline: every month end from first to last fiscal quarter in the data
        min_date = ttm_table['report_date'].min()
        max_date = ttm_table['report_date'].max()
        baseline = pd.DataFrame({
            'report_date': pd.date_range(start=min_date, end=max_date, freq='ME').astype('datetime64[us]')
        })

        total_net_income = np.zeros(len(baseline))
        total_revenue = np.zeros(len(baseline))
        count_with_data = np.zeros(len(baseline), dtype=int)

        for symbol in ni_symbols:
            ni_col = f"{symbol}{ni_suffix}"
            rev_col = f"{symbol}{rev_suffix}"
            if ni_col not in ttm_table.columns or rev_col not in ttm_table.columns:
                continue

            currency = currency_dict.get(symbol, 'USD')
            if currency == 'USD':
                fx_series = pd.Series(1.0, index=ttm_table.index)
            else:
                currency_df = self.currency(symbol=currency + '=X')
                currency_df['report_date'] = pd.to_datetime(currency_df['report_date']).astype('datetime64[us]')
                fx_series = pd.merge_asof(
                    ttm_table[['report_date']],
                    currency_df,
                    on='report_date',
                    direction='backward'
                )['close']

            symbol_df = pd.DataFrame({
                'report_date': ttm_table['report_date'],
                'ni_usd': (ttm_table[ni_col] / fx_series).values,
                'rev_usd': (ttm_table[rev_col] / fx_series).values,
            }).dropna()

            # Align each company's quarterly TTM data to the monthly baseline via forward-fill
            ni_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'ni_usd']],
                on='report_date', direction='backward'
            )['ni_usd']
            rev_aligned = pd.merge_asof(
                baseline, symbol_df[['report_date', 'rev_usd']],
                on='report_date', direction='backward'
            )['rev_usd']

            # Paired exclusion: only include when both TTM net income and revenue are available
            has_data = ni_aligned.notna().values & rev_aligned.notna().values
            total_net_income += np.where(has_data, ni_aligned.fillna(0).values, 0.0)
            total_revenue += np.where(has_data, rev_aligned.fillna(0).values, 0.0)
            count_with_data += has_data.astype(int)

        result = baseline.copy()
        result.insert(1, 'industry', industry)
        has_any = count_with_data > 0
        result['total_ttm_net_income'] = np.where(has_any, total_net_income, np.nan)
        result['total_ttm_revenue'] = np.where(has_any, total_revenue, np.nan)
        result['industry_net_margin'] = (result['total_ttm_net_income'] / result['total_ttm_revenue']).replace([np.inf, -np.inf], np.nan).round(4)
        result = result.dropna(subset=['total_ttm_net_income'])
        return result

    def industry_asset_turnover(self) -> pd.DataFrame:
        roa = self.industry_roa()
        net_margin = self.industry_quarterly_net_margin()

        roa['report_date'] = pd.to_datetime(roa['report_date']).astype('datetime64[us]')
        net_margin['report_date'] = pd.to_datetime(net_margin['report_date']).astype('datetime64[us]')

        # Both use the same monthly baseline, so a regular inner merge aligns perfectly
        result_df = roa[['report_date', 'industry', 'industry_roa']].merge(
            net_margin[['report_date', 'industry_net_margin']],
            on='report_date',
            how='inner'
        )

        result_df['industry_asset_turnover'] = (result_df['industry_roa'] / result_df['industry_net_margin']).replace([np.inf, -np.inf], np.nan).round(2)
        return result_df

    def _quarterly_eps_yoy_growth(self, eps_column: str, current_alias: str, prev_alias: str) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_tailing_eps)
        sql = load_sql("select_quarterly_eps_yoy_growth_by_symbol",
                       ticker = self.ticker,
                       url = url,
                       eps_column = eps_column,
                       current_alias = current_alias,
                       prev_alias = prev_alias)
        return self.duckdb_client.query(sql)

    def _calculate_yoy_growth(self, item_name: str, period_type: str, finance_type: str) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_statement)
        metric_name = item_name.replace('total_', '')  # For naming consistency in output
        ttm_filter = "AND report_date != 'TTM'" if period_type == 'quarterly' else ''

        sql = load_sql("select_metric_calculate_yoy_growth_by_symbol",
                       ticker = self.ticker,
                       url = url,
                       metric_name = metric_name,
                       item_name = item_name,
                       period_type = period_type,
                       finance_type = finance_type,
                       ttm_filter = ttm_filter)
        return self.duckdb_client.query(sql)

    def _revenue_by_breakdown(self, breakdown_type: str) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(stock_revenue_breakdown)
        sql = load_sql(
            "select_revenue_breakdown_by_symbol",
            ticker = self.ticker,
            url = url,
            breakdown_type = breakdown_type)
        data = self.duckdb_client.query(sql)
        df_wide = data.pivot(index=['report_date'], columns='item_name', values='item_value').reset_index()
        df_wide.columns.name = None
        df_wide = df_wide.fillna(0)
        df_wide.insert(0, 'symbol', self.ticker)
        return df_wide

    def _generate_margin(self, margin_type: str, period_type: str, numerator_item: str,
                         margin_column: str) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path('stock_statement')
        ttm_filter = "AND report_date != 'TTM'" if period_type == 'quarterly' else ""
        finance_type_filter = \
            "AND finance_type = 'income_statement'" if margin_type in ['gross', 'operating', 'net', 'ebitda'] \
            else "AND finance_type in ('income_statement', 'cash_flow')" if margin_type == 'fcf' \
            else ""
        sql = load_sql("select_margin_for_symbol",
                       ticker = self.ticker,
                       url = url,
                       numerator_item = numerator_item,
                       margin_column = margin_column,
                       period_type = period_type,
                       ttm_filter = ttm_filter,
                       finance_type_filter = finance_type_filter)
        return self.duckdb_client.query(sql)

    def _query_data(self, table_name: str) -> pd.DataFrame:
        return self._query_data2(table_name, self.ticker)

    def _query_data2(self, table_name: str, ticker: str) -> pd.DataFrame:
        url = self.huggingface_client.get_url_path(table_name)
        sql = load_sql(
            "select_all_by_symbol",
                        ticker = ticker,
                        url = url)
        return self.duckdb_client.query(sql)

    def _statement(self, finance_type: str, period_type: str) -> Statement:
        url = self.huggingface_client.get_url_path(stock_statement)
        sql = load_sql("select_statement_by_symbol",
                       url=url,
                       ticker=self.ticker,
                       finance_type=finance_type,
                       period_type=period_type)
        df = self.duckdb_client.query(sql)
        stock_statements = self._dataframe_to_stock_statements(df=df)
        if finance_type == income_statement:
            template_type = income_statement_template_type(df)
            template = load_finance_template(income_statement, template_type)
            finance_values_map = self._get_finance_values_map(statements=stock_statements, finance_template=template)
            stmt = IncomeStatement(finance_template=template, income_finance_values=finance_values_map)
            printer = PrintVisitor()
            stmt.accept(printer)
            return printer.get_statement()
        elif finance_type == balance_sheet:
            template_type = balance_sheet_template_type(df)
            template = load_finance_template(balance_sheet, template_type)
            finance_values_map = self._get_finance_values_map(statements=stock_statements, finance_template=template)
            stmt = BalanceSheet(finance_template=template, income_finance_values=finance_values_map)
            printer = PrintVisitor()
            stmt.accept(printer)
            return printer.get_statement()
        elif finance_type == cash_flow:
            template_type = cash_flow_template_type(df)
            template = load_finance_template(cash_flow, template_type)
            finance_values_map = self._get_finance_values_map(statements=stock_statements, finance_template=template)
            stmt = BalanceSheet(finance_template=template, income_finance_values=finance_values_map)
            printer = PrintVisitor()
            stmt.accept(printer)
            return printer.get_statement()
        else:
            raise ValueError(f"unknown finance type: {finance_type}")

    @staticmethod
    def _dataframe_to_stock_statements(df: pd.DataFrame) -> List[StockStatement]:
        statements = []

        for _, row in df.iterrows():
            try:
                item_value = Decimal(str(row['item_value'])) if not pd.isna(row['item_value']) else None
                statement = StockStatement(
                    symbol=str(row['symbol']),
                    report_date=str(row['report_date']),
                    item_name=str(row['item_name']),
                    item_value=item_value,
                    finance_type=str(row['finance_type']),
                    period_type=str(row['period_type'])
                )
                statements.append(statement)
            except Exception as e:
                print(f"Error processing row {row}: {str(e)}")
                continue

        return statements

    @staticmethod
    def _get_finance_values_map(statements: List['StockStatement'],
                                finance_template: Dict[str, 'FinanceItem']) -> Dict[str, List['FinanceValue']]:
        finance_item_title_keys = CaseInsensitiveDict()
        parse_all_title_keys(list(finance_template.values()), finance_item_title_keys)

        finance_values = defaultdict(list)

        for statement in statements:
            period = "TTM" if statement.report_date == "TTM" else (
                "3M" if statement.period_type == "quarterly" else "12M")
            value = FinanceValue(
                finance_key=statement.item_name,
                report_date=statement.report_date,
                report_value=statement.item_value,
                period_type=period
            )
            finance_values[statement.item_name].append(value)

        final_map = CaseInsensitiveDict()

        for title, values in finance_values.items():
            key = finance_item_title_keys.get(title)
            if key is not None:
                final_map[key] = values

        return final_map

    def download_data_performance(self) -> str:
        res = f"-------------- Download Data Performance ---------------"
        res += f"\n"
        res += self.duckdb_client.query(
            "SELECT * FROM cache_httpfs_cache_access_info_query()"
        ).to_string()
        res += f"\n"
        res += f"--------------------------------------------------------"
        return res